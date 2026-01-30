"""
BOBWEIR Pharmaceuticals - Sample Financial Model Generator

Creates a realistic specialty pharma financial model for testing the Excel Model Auditor.
Includes intentional issues for the auditor to detect:
- A few hard-coded plugs in projection years
- One balance sheet imbalance (small)
- External link reference (simulated)
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# --- STYLE DEFINITIONS ---
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
section_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Yellow for inputs
number_format = '#,##0'
pct_format = '0.0%'
currency_format = '"$"#,##0'

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_bobweir_model():
    """Creates the full BOBWEIR financial model."""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create sheets
    create_cover_sheet(wb)
    create_assumptions_sheet(wb)
    create_revenue_sheet(wb)
    create_income_statement(wb)
    create_balance_sheet(wb)
    create_cash_flow(wb)
    create_dcf_sheet(wb)

    return wb


def create_cover_sheet(wb):
    """Cover page with model info."""
    ws = wb.create_sheet("Cover")

    ws['B3'] = "BOBWEIR PHARMACEUTICALS, INC."
    ws['B3'].font = Font(bold=True, size=24, color="1F4E78")

    ws['B5'] = "Integrated Financial Model"
    ws['B5'].font = Font(size=16, italic=True)

    ws['B8'] = "Model Overview"
    ws['B8'].font = Font(bold=True, size=14)

    ws['B10'] = "Company Description:"
    ws['C10'] = "Specialty pharmaceutical company focused on rare diseases and oncology"

    ws['B11'] = "Headquarters:"
    ws['C11'] = "San Francisco, CA"

    ws['B12'] = "Founded:"
    ws['C12'] = 2008

    ws['B13'] = "Employees:"
    ws['C13'] = 2850

    ws['B15'] = "Model Details"
    ws['B15'].font = Font(bold=True, size=14)

    ws['B17'] = "Historical Period:"
    ws['C17'] = "2021A - 2023A"

    ws['B18'] = "Projection Period:"
    ws['C18'] = "2024E - 2028E"

    ws['B19'] = "Last Updated:"
    ws['C19'] = "January 2024"

    ws['B21'] = "Product Portfolio"
    ws['B21'].font = Font(bold=True, size=14)

    products = [
        ("Neurex", "Rare neurological disorder", "$850M", "2019"),
        ("Oncovir", "Solid tumor oncology", "$720M", "2020"),
        ("Hemaguard", "Rare blood disorder", "$540M", "2018"),
        ("Cardioshield", "Cardiovascular", "$420M", "2021"),
        ("Dermaclear", "Dermatology", "$310M", "2022"),
        ("Respiron", "Respiratory", "$160M", "2023"),
    ]

    ws['B23'] = "Product"
    ws['C23'] = "Indication"
    ws['D23'] = "2023 Revenue"
    ws['E23'] = "Launch Year"
    for cell in [ws['B23'], ws['C23'], ws['D23'], ws['E23']]:
        cell.font = header_font
        cell.fill = header_fill

    for i, (name, indication, rev, launch) in enumerate(products, start=24):
        ws[f'B{i}'] = name
        ws[f'C{i}'] = indication
        ws[f'D{i}'] = rev
        ws[f'E{i}'] = launch

    # Set column widths
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15


def create_assumptions_sheet(wb):
    """Key assumptions driving the model."""
    ws = wb.create_sheet("Assumptions")

    # Headers
    headers = ["Assumption", "Units", "2021A", "2022A", "2023A", "2024E", "2025E", "2026E", "2027E", "2028E"]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Revenue Growth Assumptions
    ws['B4'] = "REVENUE ASSUMPTIONS"
    ws['B4'].font = Font(bold=True)
    ws['B4'].fill = section_fill

    assumptions = [
        ("Neurex Growth Rate", "%", 0.15, 0.12, 0.08, 0.06, 0.04, 0.03, 0.02, 0.01),
        ("Oncovir Growth Rate", "%", 0.25, 0.22, 0.18, 0.15, 0.12, 0.10, 0.08, 0.06),
        ("Hemaguard Growth Rate", "%", 0.10, 0.08, 0.05, 0.03, 0.02, 0.01, 0.00, -0.02),
        ("Cardioshield Growth Rate", "%", None, 0.80, 0.35, 0.25, 0.18, 0.12, 0.08, 0.05),
        ("Dermaclear Growth Rate", "%", None, None, 1.50, 0.40, 0.30, 0.22, 0.15, 0.10),
        ("Respiron Growth Rate", "%", None, None, None, 0.60, 0.45, 0.35, 0.25, 0.18),
    ]

    for i, (name, units, *values) in enumerate(assumptions, start=5):
        ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=3, value=units)
        for j, val in enumerate(values, start=4):
            cell = ws.cell(row=i, column=j, value=val)
            if val is not None:
                cell.number_format = pct_format
            if j >= 7:  # Projection years
                cell.fill = input_fill

    # Margin Assumptions
    ws['B12'] = "MARGIN ASSUMPTIONS"
    ws['B12'].font = Font(bold=True)
    ws['B12'].fill = section_fill

    margins = [
        ("Gross Margin", "%", 0.72, 0.73, 0.74, 0.745, 0.75, 0.755, 0.76, 0.765),
        ("R&D % of Revenue", "%", 0.18, 0.17, 0.16, 0.155, 0.15, 0.145, 0.14, 0.135),
        ("SG&A % of Revenue", "%", 0.28, 0.27, 0.26, 0.25, 0.24, 0.23, 0.22, 0.21),
        ("D&A % of Revenue", "%", 0.04, 0.04, 0.04, 0.04, 0.04, 0.04, 0.04, 0.04),
    ]

    for i, (name, units, *values) in enumerate(margins, start=13):
        ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=3, value=units)
        for j, val in enumerate(values, start=4):
            cell = ws.cell(row=i, column=j, value=val)
            cell.number_format = pct_format
            if j >= 7:
                cell.fill = input_fill

    # Balance Sheet Assumptions
    ws['B19'] = "BALANCE SHEET ASSUMPTIONS"
    ws['B19'].font = Font(bold=True)
    ws['B19'].fill = section_fill

    bs_assumptions = [
        ("Days Sales Outstanding", "days", 55, 52, 50, 48, 46, 45, 44, 43),
        ("Days Inventory", "days", 90, 88, 85, 82, 80, 78, 76, 75),
        ("Days Payable", "days", 45, 48, 50, 52, 54, 55, 56, 57),
        ("CapEx % of Revenue", "%", 0.05, 0.05, 0.05, 0.045, 0.045, 0.04, 0.04, 0.04),
    ]

    for i, (name, units, *values) in enumerate(bs_assumptions, start=20):
        ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=3, value=units)
        for j, val in enumerate(values, start=4):
            cell = ws.cell(row=i, column=j, value=val)
            if "%" in units:
                cell.number_format = pct_format
            if j >= 7:
                cell.fill = input_fill

    # Tax & Other
    ws['B26'] = "TAX & OTHER"
    ws['B26'].font = Font(bold=True)
    ws['B26'].fill = section_fill

    ws.cell(row=27, column=2, value="Effective Tax Rate")
    ws.cell(row=27, column=3, value="%")
    for j, val in enumerate([0.21, 0.21, 0.21, 0.21, 0.21, 0.21, 0.21, 0.21], start=4):
        cell = ws.cell(row=27, column=j, value=val)
        cell.number_format = pct_format

    # Column widths
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 8
    for col in range(4, 12):
        ws.column_dimensions[get_column_letter(col)].width = 12


def create_revenue_sheet(wb):
    """Detailed revenue by product."""
    ws = wb.create_sheet("Revenue")

    # Headers
    headers = ["", "", "2021A", "2022A", "2023A", "2024E", "2025E", "2026E", "2027E", "2028E"]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    ws['B4'] = "PRODUCT REVENUE ($M)"
    ws['B4'].font = Font(bold=True)
    ws['B4'].fill = section_fill

    # Base revenues for 2021
    products = {
        'Neurex': {'base': 650, 'row': 5},
        'Oncovir': {'base': 380, 'row': 6},
        'Hemaguard': {'base': 480, 'row': 7},
        'Cardioshield': {'base': 0, 'row': 8},  # Launched 2021
        'Dermaclear': {'base': 0, 'row': 9},     # Launched 2022
        'Respiron': {'base': 0, 'row': 10},      # Launched 2023
    }

    # 2021A values
    revenue_data = {
        'Neurex': [650, 728, 786, 833, 866, 892, 910, 919],
        'Oncovir': [380, 475, 561, 645, 722, 794, 858, 909],
        'Hemaguard': [480, 518, 544, 560, 571, 577, 577, 565],
        'Cardioshield': [180, 324, 437, 546, 644, 721, 779, 818],
        'Dermaclear': [0, 85, 212, 297, 386, 471, 542, 596],
        'Respiron': [0, 0, 160, 256, 371, 501, 626, 739],
    }

    for product, data in revenue_data.items():
        row = products[product]['row']
        ws.cell(row=row, column=2, value=product)
        for col, val in enumerate(data, start=4):
            cell = ws.cell(row=row, column=col, value=val)
            cell.number_format = number_format
            # Add formulas for projection years (with one intentional plug)
            if col >= 7:  # 2024E onwards
                if product == 'Neurex' and col == 8:  # INTENTIONAL PLUG in 2025
                    cell.value = 866  # Hard-coded instead of formula
                else:
                    # Reference to growth rate in Assumptions
                    prev_col = get_column_letter(col - 1)
                    growth_row = {'Neurex': 5, 'Oncovir': 6, 'Hemaguard': 7,
                                  'Cardioshield': 8, 'Dermaclear': 9, 'Respiron': 10}[product]
                    cell.value = f"={prev_col}{row}*(1+Assumptions!{get_column_letter(col)}{growth_row})"

    # Total Revenue row
    ws['B12'] = "Total Revenue"
    ws['B12'].font = Font(bold=True)
    for col in range(4, 12):
        cell = ws.cell(row=12, column=col)
        col_letter = get_column_letter(col)
        cell.value = f"=SUM({col_letter}5:{col_letter}10)"
        cell.font = Font(bold=True)
        cell.number_format = number_format

    # YoY Growth row
    ws['B14'] = "YoY Growth %"
    ws.cell(row=14, column=4, value="N/A")
    for col in range(5, 12):
        cell = ws.cell(row=14, column=col)
        prev_col = get_column_letter(col - 1)
        col_letter = get_column_letter(col)
        cell.value = f"={col_letter}12/{prev_col}12-1"
        cell.number_format = pct_format

    # Column widths
    ws.column_dimensions['B'].width = 18
    for col in range(4, 12):
        ws.column_dimensions[get_column_letter(col)].width = 12


def create_income_statement(wb):
    """Income statement with formulas linking to other sheets."""
    ws = wb.create_sheet("IS")

    # Headers
    headers = ["", "", "2021A", "2022A", "2023A", "2024E", "2025E", "2026E", "2027E", "2028E"]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    ws['B4'] = "INCOME STATEMENT ($M)"
    ws['B4'].font = Font(bold=True)
    ws['B4'].fill = section_fill

    # Line items with their row numbers
    line_items = [
        ("Revenue", 6, "=Revenue!{col}12"),
        ("Cost of Goods Sold", 7, "=-{col}6*(1-Assumptions!{col}13)"),
        ("Gross Profit", 8, "={col}6+{col}7"),
        ("", 9, None),
        ("R&D Expense", 10, "=-{col}6*Assumptions!{col}14"),
        ("SG&A Expense", 11, "=-{col}6*Assumptions!{col}15"),
        ("D&A Expense", 12, "=-{col}6*Assumptions!{col}16"),
        ("Operating Income (EBIT)", 13, "={col}8+{col}10+{col}11+{col}12"),
        ("", 14, None),
        ("Interest Expense", 15, None),  # Will be hard-coded
        ("Interest Income", 16, None),
        ("Pre-Tax Income", 17, "={col}13+{col}15+{col}16"),
        ("", 18, None),
        ("Income Tax Expense", 19, "=-{col}17*Assumptions!{col}27"),
        ("Net Income", 20, "={col}17+{col}19"),
    ]

    # Historical interest values
    interest_expense = [-45, -42, -40, -38, -36, -34, -32, -30]
    interest_income = [5, 8, 12, 15, 18, 22, 26, 30]

    for item_name, row, formula_template in line_items:
        ws.cell(row=row, column=2, value=item_name)
        if item_name in ["Gross Profit", "Operating Income (EBIT)", "Pre-Tax Income", "Net Income"]:
            ws.cell(row=row, column=2).font = Font(bold=True)

        for col_idx in range(4, 12):
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(row=row, column=col_idx)

            if item_name == "Interest Expense":
                cell.value = interest_expense[col_idx - 4]
            elif item_name == "Interest Income":
                cell.value = interest_income[col_idx - 4]
            elif formula_template:
                cell.value = formula_template.format(col=col_letter)

            if item_name:
                cell.number_format = number_format

    # Margins section
    ws['B22'] = "MARGINS"
    ws['B22'].font = Font(bold=True)
    ws['B22'].fill = section_fill

    margin_items = [
        ("Gross Margin", 23, "={col}8/{col}6"),
        ("R&D Margin", 24, "=-{col}10/{col}6"),
        ("SG&A Margin", 25, "=-{col}11/{col}6"),
        ("Operating Margin", 26, "={col}13/{col}6"),
        ("Net Margin", 27, "={col}20/{col}6"),
    ]

    for item_name, row, formula_template in margin_items:
        ws.cell(row=row, column=2, value=item_name)
        for col_idx in range(4, 12):
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(row=row, column=col_idx)
            cell.value = formula_template.format(col=col_letter)
            cell.number_format = pct_format

    # Column widths
    ws.column_dimensions['B'].width = 25
    for col in range(4, 12):
        ws.column_dimensions[get_column_letter(col)].width = 12


def create_balance_sheet(wb):
    """Balance sheet with working capital formulas."""
    ws = wb.create_sheet("BS")

    # Headers
    headers = ["", "", "2021A", "2022A", "2023A", "2024E", "2025E", "2026E", "2027E", "2028E"]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    ws['B4'] = "BALANCE SHEET ($M)"
    ws['B4'].font = Font(bold=True)
    ws['B4'].fill = section_fill

    # ASSETS
    ws['B6'] = "ASSETS"
    ws['B6'].font = Font(bold=True)

    # Current Assets - using historical values then formulas
    cash_values = [450, 580, 720, None, None, None, None, None]  # Will link to CF
    ar_base = [260, 295, 342, None, None, None, None, None]  # DSO formula
    inv_base = [380, 420, 465, None, None, None, None, None]  # DIO formula

    # Cash row
    ws.cell(row=7, column=2, value="Cash & Equivalents")
    for col_idx in range(4, 12):
        cell = ws.cell(row=7, column=col_idx)
        if col_idx < 7:  # Historical
            cell.value = cash_values[col_idx - 4]
        else:  # Projection - link to prior year + CF
            prev_col = get_column_letter(col_idx - 1)
            cell.value = f"={prev_col}7+CF!{get_column_letter(col_idx)}28"
        cell.number_format = number_format

    # Accounts Receivable
    ws.cell(row=8, column=2, value="Accounts Receivable")
    for col_idx in range(4, 12):
        cell = ws.cell(row=8, column=col_idx)
        col_letter = get_column_letter(col_idx)
        if col_idx < 7:
            cell.value = ar_base[col_idx - 4]
        else:
            cell.value = f"=Revenue!{col_letter}12*Assumptions!{col_letter}20/365"
        cell.number_format = number_format

    # Inventory
    ws.cell(row=9, column=2, value="Inventory")
    for col_idx in range(4, 12):
        cell = ws.cell(row=9, column=col_idx)
        col_letter = get_column_letter(col_idx)
        if col_idx < 7:
            cell.value = inv_base[col_idx - 4]
        else:
            cell.value = f"=-IS!{col_letter}7*Assumptions!{col_letter}21/365"
        cell.number_format = number_format

    # Other Current Assets
    ws.cell(row=10, column=2, value="Other Current Assets")
    for col_idx in range(4, 12):
        cell = ws.cell(row=10, column=col_idx)
        cell.value = 50 + (col_idx - 4) * 5
        cell.number_format = number_format

    # Total Current Assets
    ws.cell(row=11, column=2, value="Total Current Assets")
    ws.cell(row=11, column=2).font = Font(bold=True)
    for col_idx in range(4, 12):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=11, column=col_idx)
        cell.value = f"=SUM({col_letter}7:{col_letter}10)"
        cell.font = Font(bold=True)
        cell.number_format = number_format

    # Non-Current Assets
    ws.cell(row=13, column=2, value="PP&E, Net")
    ppe_base = [320, 350, 385, None, None, None, None, None]
    for col_idx in range(4, 12):
        cell = ws.cell(row=13, column=col_idx)
        col_letter = get_column_letter(col_idx)
        if col_idx < 7:
            cell.value = ppe_base[col_idx - 4]
        else:
            prev_col = get_column_letter(col_idx - 1)
            cell.value = f"={prev_col}13+Revenue!{col_letter}12*Assumptions!{col_letter}23+IS!{col_letter}12"
        cell.number_format = number_format

    # Intangibles
    ws.cell(row=14, column=2, value="Intangible Assets")
    for col_idx in range(4, 12):
        cell = ws.cell(row=14, column=col_idx)
        cell.value = 850 - (col_idx - 4) * 30  # Amortizing
        cell.number_format = number_format

    # Goodwill
    ws.cell(row=15, column=2, value="Goodwill")
    for col_idx in range(4, 12):
        cell = ws.cell(row=15, column=col_idx)
        cell.value = 420
        cell.number_format = number_format

    # Total Non-Current Assets
    ws.cell(row=16, column=2, value="Total Non-Current Assets")
    ws.cell(row=16, column=2).font = Font(bold=True)
    for col_idx in range(4, 12):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=16, column=col_idx)
        cell.value = f"=SUM({col_letter}13:{col_letter}15)"
        cell.font = Font(bold=True)
        cell.number_format = number_format

    # TOTAL ASSETS
    ws.cell(row=18, column=2, value="TOTAL ASSETS")
    ws.cell(row=18, column=2).font = Font(bold=True)
    for col_idx in range(4, 12):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=18, column=col_idx)
        cell.value = f"={col_letter}11+{col_letter}16"
        cell.font = Font(bold=True)
        cell.number_format = number_format

    # LIABILITIES
    ws['B20'] = "LIABILITIES"
    ws['B20'].font = Font(bold=True)

    # Accounts Payable
    ws.cell(row=21, column=2, value="Accounts Payable")
    ap_base = [170, 195, 220, None, None, None, None, None]
    for col_idx in range(4, 12):
        cell = ws.cell(row=21, column=col_idx)
        col_letter = get_column_letter(col_idx)
        if col_idx < 7:
            cell.value = ap_base[col_idx - 4]
        else:
            cell.value = f"=-IS!{col_letter}7*Assumptions!{col_letter}22/365"
        cell.number_format = number_format

    # Accrued Expenses
    ws.cell(row=22, column=2, value="Accrued Expenses")
    for col_idx in range(4, 12):
        cell = ws.cell(row=22, column=col_idx)
        cell.value = 120 + (col_idx - 4) * 10
        cell.number_format = number_format

    # Total Current Liabilities
    ws.cell(row=23, column=2, value="Total Current Liabilities")
    ws.cell(row=23, column=2).font = Font(bold=True)
    for col_idx in range(4, 12):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=23, column=col_idx)
        cell.value = f"=SUM({col_letter}21:{col_letter}22)"
        cell.font = Font(bold=True)
        cell.number_format = number_format

    # Long-term Debt
    ws.cell(row=25, column=2, value="Long-term Debt")
    debt_values = [600, 550, 500, 450, 400, 350, 300, 250]
    for col_idx in range(4, 12):
        cell = ws.cell(row=25, column=col_idx)
        cell.value = debt_values[col_idx - 4]
        cell.number_format = number_format

    # Other LT Liabilities
    ws.cell(row=26, column=2, value="Other Long-term Liabilities")
    for col_idx in range(4, 12):
        cell = ws.cell(row=26, column=col_idx)
        cell.value = 80
        cell.number_format = number_format

    # Total Long-term Liabilities
    ws.cell(row=27, column=2, value="Total Long-term Liabilities")
    ws.cell(row=27, column=2).font = Font(bold=True)
    for col_idx in range(4, 12):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=27, column=col_idx)
        cell.value = f"=SUM({col_letter}25:{col_letter}26)"
        cell.font = Font(bold=True)
        cell.number_format = number_format

    # TOTAL LIABILITIES
    ws.cell(row=29, column=2, value="TOTAL LIABILITIES")
    ws.cell(row=29, column=2).font = Font(bold=True)
    for col_idx in range(4, 12):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=29, column=col_idx)
        cell.value = f"={col_letter}23+{col_letter}27"
        cell.font = Font(bold=True)
        cell.number_format = number_format

    # EQUITY
    ws['B31'] = "EQUITY"
    ws['B31'].font = Font(bold=True)

    # Common Stock
    ws.cell(row=32, column=2, value="Common Stock")
    for col_idx in range(4, 12):
        cell = ws.cell(row=32, column=col_idx)
        cell.value = 200
        cell.number_format = number_format

    # Retained Earnings
    ws.cell(row=33, column=2, value="Retained Earnings")
    re_base = 890  # Starting RE
    for col_idx in range(4, 12):
        cell = ws.cell(row=33, column=col_idx)
        col_letter = get_column_letter(col_idx)
        if col_idx == 4:
            cell.value = re_base
        else:
            prev_col = get_column_letter(col_idx - 1)
            cell.value = f"={prev_col}33+IS!{col_letter}20"
        cell.number_format = number_format

    # TOTAL EQUITY
    ws.cell(row=34, column=2, value="TOTAL EQUITY")
    ws.cell(row=34, column=2).font = Font(bold=True)
    for col_idx in range(4, 12):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=34, column=col_idx)
        cell.value = f"={col_letter}32+{col_letter}33"
        cell.font = Font(bold=True)
        cell.number_format = number_format

    # TOTAL LIABILITIES & EQUITY
    ws.cell(row=36, column=2, value="TOTAL LIABILITIES & EQUITY")
    ws.cell(row=36, column=2).font = Font(bold=True)
    for col_idx in range(4, 12):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=36, column=col_idx)
        cell.value = f"={col_letter}29+{col_letter}34"
        cell.font = Font(bold=True)
        cell.number_format = number_format

    # Balance Check row
    ws.cell(row=38, column=2, value="Balance Check (should be 0)")
    ws.cell(row=38, column=2).font = Font(italic=True)
    for col_idx in range(4, 12):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=38, column=col_idx)
        cell.value = f"={col_letter}18-{col_letter}36"
        cell.number_format = number_format

    # Column widths
    ws.column_dimensions['B'].width = 28
    for col in range(4, 12):
        ws.column_dimensions[get_column_letter(col)].width = 12


def create_cash_flow(wb):
    """Cash flow statement."""
    ws = wb.create_sheet("CF")

    # Headers
    headers = ["", "", "2021A", "2022A", "2023A", "2024E", "2025E", "2026E", "2027E", "2028E"]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    ws['B4'] = "CASH FLOW STATEMENT ($M)"
    ws['B4'].font = Font(bold=True)
    ws['B4'].fill = section_fill

    # Operating Activities
    ws['B6'] = "OPERATING ACTIVITIES"
    ws['B6'].font = Font(bold=True)

    cf_items = [
        ("Net Income", 7, "=IS!{col}20"),
        ("D&A (add back)", 8, "=-IS!{col}12"),
        ("Change in Working Capital", 9, None),  # Complex formula
        ("Cash from Operations", 11, "={col}7+{col}8+{col}9"),
    ]

    for item_name, row, formula_template in cf_items:
        ws.cell(row=row, column=2, value=item_name)
        if "Cash from" in item_name:
            ws.cell(row=row, column=2).font = Font(bold=True)

        for col_idx in range(4, 12):
            col_letter = get_column_letter(col_idx)
            prev_col = get_column_letter(col_idx - 1) if col_idx > 4 else None
            cell = ws.cell(row=row, column=col_idx)

            if item_name == "Change in Working Capital":
                if col_idx == 4:
                    cell.value = -25  # Base year
                else:
                    # (Change in AR + Change in Inv) - Change in AP
                    cell.value = f"=-(BS!{col_letter}8-BS!{prev_col}8)-(BS!{col_letter}9-BS!{prev_col}9)+(BS!{col_letter}21-BS!{prev_col}21)"
            elif formula_template:
                cell.value = formula_template.format(col=col_letter)

            cell.number_format = number_format

    # Investing Activities
    ws['B13'] = "INVESTING ACTIVITIES"
    ws['B13'].font = Font(bold=True)

    ws.cell(row=14, column=2, value="Capital Expenditures")
    for col_idx in range(4, 12):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=14, column=col_idx)
        cell.value = f"=-Revenue!{col_letter}12*Assumptions!{col_letter}23"
        cell.number_format = number_format

    ws.cell(row=15, column=2, value="Cash from Investing")
    ws.cell(row=15, column=2).font = Font(bold=True)
    for col_idx in range(4, 12):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=15, column=col_idx)
        cell.value = f"={col_letter}14"
        cell.font = Font(bold=True)
        cell.number_format = number_format

    # Financing Activities
    ws['B17'] = "FINANCING ACTIVITIES"
    ws['B17'].font = Font(bold=True)

    ws.cell(row=18, column=2, value="Debt Repayment")
    for col_idx in range(4, 12):
        cell = ws.cell(row=18, column=col_idx)
        if col_idx == 4:
            cell.value = 0
        else:
            prev_col = get_column_letter(col_idx - 1)
            col_letter = get_column_letter(col_idx)
            cell.value = f"=BS!{col_letter}25-BS!{prev_col}25"
        cell.number_format = number_format

    ws.cell(row=19, column=2, value="Dividends Paid")
    for col_idx in range(4, 12):
        cell = ws.cell(row=19, column=col_idx)
        cell.value = 0  # No dividends
        cell.number_format = number_format

    ws.cell(row=20, column=2, value="Cash from Financing")
    ws.cell(row=20, column=2).font = Font(bold=True)
    for col_idx in range(4, 12):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=20, column=col_idx)
        cell.value = f"={col_letter}18+{col_letter}19"
        cell.font = Font(bold=True)
        cell.number_format = number_format

    # Summary
    ws['B22'] = "SUMMARY"
    ws['B22'].font = Font(bold=True)
    ws['B22'].fill = section_fill

    ws.cell(row=24, column=2, value="Beginning Cash")
    for col_idx in range(4, 12):
        cell = ws.cell(row=24, column=col_idx)
        if col_idx == 4:
            cell.value = 380
        else:
            prev_col = get_column_letter(col_idx - 1)
            cell.value = f"={prev_col}27"
        cell.number_format = number_format

    ws.cell(row=25, column=2, value="Net Change in Cash")
    for col_idx in range(4, 12):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=25, column=col_idx)
        cell.value = f"={col_letter}11+{col_letter}15+{col_letter}20"
        cell.number_format = number_format

    ws.cell(row=27, column=2, value="Ending Cash")
    ws.cell(row=27, column=2).font = Font(bold=True)
    for col_idx in range(4, 12):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=27, column=col_idx)
        cell.value = f"={col_letter}24+{col_letter}25"
        cell.font = Font(bold=True)
        cell.number_format = number_format

    # Cash check vs BS
    ws.cell(row=28, column=2, value="Change in Cash (for BS)")
    for col_idx in range(4, 12):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=28, column=col_idx)
        cell.value = f"={col_letter}25"
        cell.number_format = number_format

    # Column widths
    ws.column_dimensions['B'].width = 25
    for col in range(4, 12):
        ws.column_dimensions[get_column_letter(col)].width = 12


def create_dcf_sheet(wb):
    """DCF valuation."""
    ws = wb.create_sheet("DCF")

    # Headers
    headers = ["", "", "2024E", "2025E", "2026E", "2027E", "2028E", "Terminal"]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    ws['B4'] = "DCF VALUATION ($M)"
    ws['B4'].font = Font(bold=True)
    ws['B4'].fill = section_fill

    # Assumptions
    ws['B6'] = "KEY ASSUMPTIONS"
    ws['B6'].font = Font(bold=True)

    ws.cell(row=7, column=2, value="WACC")
    ws.cell(row=7, column=4, value=0.10)
    ws.cell(row=7, column=4).number_format = pct_format
    ws.cell(row=7, column=4).fill = input_fill

    ws.cell(row=8, column=2, value="Terminal Growth Rate")
    ws.cell(row=8, column=4, value=0.025)
    ws.cell(row=8, column=4).number_format = pct_format
    ws.cell(row=8, column=4).fill = input_fill

    # Free Cash Flow Build
    ws['B11'] = "FREE CASH FLOW"
    ws['B11'].font = Font(bold=True)
    ws['B11'].fill = section_fill

    fcf_items = [
        ("EBIT", 12, "=IS!{col}13"),
        ("Less: Taxes", 13, "=-{col}12*Assumptions!{col}27"),
        ("EBIAT", 14, "={col}12+{col}13"),
        ("Plus: D&A", 15, "=-IS!{col}12"),
        ("Less: CapEx", 16, "=CF!{col}14"),
        ("Less: Change in NWC", 17, "=CF!{col}9"),
        ("Unlevered Free Cash Flow", 18, "={col}14+{col}15+{col}16+{col}17"),
    ]

    # Map DCF columns to IS/CF columns (DCF starts at 2024)
    dcf_to_is_col = {4: 7, 5: 8, 6: 9, 7: 10, 8: 11}  # DCF col -> IS col

    for item_name, row, formula_template in fcf_items:
        ws.cell(row=row, column=2, value=item_name)
        if item_name in ["EBIAT", "Unlevered Free Cash Flow"]:
            ws.cell(row=row, column=2).font = Font(bold=True)

        for dcf_col in range(4, 9):
            is_col = dcf_to_is_col[dcf_col]
            is_col_letter = get_column_letter(is_col)
            dcf_col_letter = get_column_letter(dcf_col)
            cell = ws.cell(row=row, column=dcf_col)

            if "{col}" in str(formula_template):
                # Replace with IS column for references to other sheets
                if "IS!" in formula_template or "CF!" in formula_template or "Assumptions!" in formula_template:
                    cell.value = formula_template.format(col=is_col_letter)
                else:
                    cell.value = formula_template.format(col=dcf_col_letter)

            cell.number_format = number_format

    # Terminal Value
    ws.cell(row=18, column=9, value="=H18*(1+$D$8)/($D$7-$D$8)")
    ws.cell(row=18, column=9).number_format = number_format

    # Discount factors and PV
    ws['B20'] = "DISCOUNTING"
    ws['B20'].font = Font(bold=True)
    ws['B20'].fill = section_fill

    ws.cell(row=21, column=2, value="Period")
    for i, dcf_col in enumerate(range(4, 10), start=1):
        ws.cell(row=21, column=dcf_col, value=i if dcf_col < 9 else 5)

    ws.cell(row=22, column=2, value="Discount Factor")
    for dcf_col in range(4, 10):
        dcf_col_letter = get_column_letter(dcf_col)
        cell = ws.cell(row=22, column=dcf_col)
        cell.value = f"=1/(1+$D$7)^{dcf_col_letter}21"
        cell.number_format = "0.000"

    ws.cell(row=23, column=2, value="PV of Cash Flow")
    ws.cell(row=23, column=2).font = Font(bold=True)
    for dcf_col in range(4, 10):
        dcf_col_letter = get_column_letter(dcf_col)
        cell = ws.cell(row=23, column=dcf_col)
        cell.value = f"={dcf_col_letter}18*{dcf_col_letter}22"
        cell.font = Font(bold=True)
        cell.number_format = number_format

    # Valuation Summary
    ws['B26'] = "VALUATION SUMMARY"
    ws['B26'].font = Font(bold=True)
    ws['B26'].fill = section_fill

    ws.cell(row=28, column=2, value="PV of Forecast Period FCF")
    ws.cell(row=28, column=4, value="=SUM(D23:H23)")
    ws.cell(row=28, column=4).number_format = number_format

    ws.cell(row=29, column=2, value="PV of Terminal Value")
    ws.cell(row=29, column=4, value="=I23")
    ws.cell(row=29, column=4).number_format = number_format

    ws.cell(row=30, column=2, value="Enterprise Value")
    ws.cell(row=30, column=2).font = Font(bold=True)
    ws.cell(row=30, column=4, value="=D28+D29")
    ws.cell(row=30, column=4).font = Font(bold=True)
    ws.cell(row=30, column=4).number_format = number_format

    ws.cell(row=32, column=2, value="Less: Net Debt (2023)")
    ws.cell(row=32, column=4, value="=BS!F25-BS!F7")  # Debt - Cash from 2023
    ws.cell(row=32, column=4).number_format = number_format

    ws.cell(row=33, column=2, value="Equity Value")
    ws.cell(row=33, column=2).font = Font(bold=True)
    ws.cell(row=33, column=4, value="=D30-D32")
    ws.cell(row=33, column=4).font = Font(bold=True)
    ws.cell(row=33, column=4).number_format = number_format

    ws.cell(row=35, column=2, value="Shares Outstanding (M)")
    ws.cell(row=35, column=4, value=150)
    ws.cell(row=35, column=4).fill = input_fill

    ws.cell(row=36, column=2, value="Implied Share Price")
    ws.cell(row=36, column=2).font = Font(bold=True, size=12)
    ws.cell(row=36, column=4, value="=D33/D35")
    ws.cell(row=36, column=4).font = Font(bold=True, size=12)
    ws.cell(row=36, column=4).number_format = '"$"#,##0.00'

    # Column widths
    ws.column_dimensions['B'].width = 28
    for col in range(4, 10):
        ws.column_dimensions[get_column_letter(col)].width = 12


def main():
    """Generate and save the model."""
    print("Creating BOBWEIR Pharmaceuticals financial model...")

    wb = create_bobweir_model()

    # Save to sample_models directory
    output_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    output_path = os.path.join(output_dir, "sample_models", "BOBWEIR_Model.xlsx")

    # Ensure directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb.save(output_path)
    print(f"Model saved to: {output_path}")
    print("\nModel includes:")
    print("  - Cover sheet with company overview")
    print("  - Assumptions (growth rates, margins, working capital)")
    print("  - Revenue by product (6 products)")
    print("  - Income Statement with linked formulas")
    print("  - Balance Sheet with working capital calculations")
    print("  - Cash Flow Statement")
    print("  - DCF Valuation")
    print("\nIntentional issues for auditor testing:")
    print("  - 1 hard-coded plug in Revenue sheet (Neurex 2025E)")
    print("  - Formula linkages that can be traced")


if __name__ == "__main__":
    main()
