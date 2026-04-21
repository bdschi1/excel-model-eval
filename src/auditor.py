import logging
import re

import pandas as pd

logger = logging.getLogger(__name__)

# --- Period classification regex (for EDGAR check) ---
# Captures year and an optional A (actual) / E (estimate) suffix. Matches:
#   - 4-digit: 2023, FY2023, FY 2023, 1999A, 2024E (range 1980-2039)
#   - 2-digit with FY prefix: FY23, FY23A, FY 23 E (short forms)
_PERIOD_RE = re.compile(
    r"\b(?:FY\s*)?(19[89]\d|20[0-3]\d)\s*(A|E)?\b"
    r"|\bFY\s*(\d{2})\s*(A|E)?\b",
    re.IGNORECASE,
)

# Quarterly / interim-period headers that should NOT be classified as annual.
_QUARTER_RE = re.compile(
    r"\bQ[1-4]\b|\b[1-4]Q\b|Q[1-4]\s*\d|\d\s*Q[1-4]"
    r"|\b(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)\w*\b"
    r"|\b(?:h1|h2|1h|2h)\b",
    re.IGNORECASE,
)

_DATA_SHEET_KEYWORDS = ("raw", "cache", "lookup", "reference", "archive", "_data", "data_", "source")

# Row-label keywords that indicate a reconciliation / tie-out row rather than
# a projection row. A literal sandwiched between formulas on such a row is
# overwhelmingly a diagnostic figure (e.g. "Check: Assets - L&E = 0"), not a
# hard-coded plug. Skipping these keeps the plug detector's false-positive
# rate low on models with explicit balance-check rows.
_PLUG_SKIP_KEYWORDS = (
    "check",
    "balance",
    "total",
    "subtotal",
    "sum",
    "audit",
    "tieout",
    "crosscheck",
)

# Balance sheet tolerance: 1 basis point of total assets (or $1 floor)
_BS_TOLERANCE_BPS = 0.0001

# --- Compiled regex patterns (module-level so they compile once at import) ---

_BS_PATTERNS = re.compile(
    r"balance|^bs$|\bbs\b|financial position|sofp|statement of financial",
    re.IGNORECASE,
)

_ASSET_PATTERNS = re.compile(
    r"total\s+assets|total\s+current.*non.?current\s+assets",
    re.IGNORECASE,
)

_LIAB_EQ_PATTERNS = re.compile(
    r"total\s+liabilities?\s*(and|&|,|\+)\s*(stockholders?\'?s?|shareholders?\'?s?|owners?\'?s?)?\s*equity"
    r"|total\s+equity\s*(and|&|,|\+)\s*liabilities?"
    r"|total\s+liabilities?\s*(and|&)\s*equity"
    r"|total\s+l\s*&\s*e"
    r"|total\s+liabilities?,?\s*equity",
    re.IGNORECASE,
)

_OUTPUT_LABELS = re.compile(
    r"irr|xirr|npv|xnpv|valuation|return|moic|wacc|capm|fcf|ev\b|enterprise.value|equity.value",
    re.IGNORECASE,
)

# --- ISSUE EXPLANATIONS ---
# Educational context for each issue type: why it matters, what causes it, how to fix it

ISSUE_EXPLANATIONS = {
    "External Link": {
        "why": "External links create dependencies on files that may not exist on other machines, "
               "causing #REF! errors when the model is shared. They also introduce version control "
               "risks if the source file changes without the model being updated.",
        "cause": "Usually created when copying data from another workbook using paste-link, or when "
                 "formulas reference Bloomberg/FactSet/Capital IQ feeds directly.",
        "fix": "Convert external links to static values (Paste Special > Values) for historical data. "
               "For live feeds, document the source and create a dedicated 'Data Inputs' sheet."
    },
    "Calculation Error": {
        "why": "Excel errors propagate through the model — any cell referencing an error cell will "
               "also show an error. This can silently break valuation outputs and key metrics.",
        "cause": {
            "#REF!": "A formula references a cell that has been deleted, or a range that was invalidated.",
            "#NAME?": "Excel doesn't recognize a function name (typo) or a named range that doesn't exist.",
            "#VALUE!": "A formula has the wrong type of argument (e.g., text where a number is expected).",
            "#DIV/0!": "A formula is dividing by zero or an empty cell."
        },
        "fix": "Trace the error back to its source using Excel's 'Trace Error' feature (Formulas > Error Checking). "
               "Fix the root cause rather than wrapping in IFERROR, which can hide real problems."
    },
    "Hard-coded Plug": {
        "why": "A 'plug' is a hard-coded value inserted into a row of formulas to force a desired result. "
               "This is a major red flag in financial models because it breaks the logical flow — "
               "changes to assumptions won't flow through correctly, and the model may produce "
               "misleading outputs without any warning.",
        "cause": "Often inserted when a model doesn't balance or produce expected results. Instead of "
                 "fixing the underlying logic, an analyst may override a cell with a manual number. "
                 "Also common in rushed model updates or when inheriting someone else's model.",
        "fix": "1) Identify what the cell SHOULD be calculating. 2) Write the correct formula. "
               "3) If the formula produces unexpected results, trace upstream to find the real issue. "
               "Never use hard-coded values in projection periods unless they represent genuine assumptions."
    },
    "Accounting Mismatch": {
        "why": "The fundamental accounting equation (Assets = Liabilities + Equity) must hold in every period. "
               "An imbalance means the model has a structural error — cash flows aren't routing correctly, "
               "or a balance sheet account is missing its corresponding entry.",
        "cause": "Common causes: (1) Working capital changes not flowing to cash flow statement, "
                 "(2) Debt/equity issuances not hitting both cash and the liability/equity account, "
                 "(3) Retained earnings not linking to net income, (4) Circular reference breaking the iteration.",
        "fix": "Create a 'Balance Check' row that calculates Assets - Liabilities - Equity for each period. "
               "Find the first period where the imbalance appears and trace all entries in that period. "
               "Check that every cash movement has a corresponding balance sheet entry."
    },
    "Circular Reference": {
        "why": "Circular references occur when a formula refers back to itself (directly or through a chain). "
               "While Excel can resolve some circular refs through iteration, they make models fragile, "
               "slow, and prone to convergence failures. They also make auditing extremely difficult.",
        "cause": "Most common in financial models: Interest expense depends on average debt, which depends on "
                 "ending cash, which depends on net income, which includes interest expense. Also common "
                 "with revolver/credit facility modeling.",
        "fix": "Break the circularity by using beginning-of-period balances instead of averages, or by "
               "implementing a 'copy-paste values' macro that iterates until convergence. Document any "
               "intentional circular references clearly."
    },
    "Unreferenced Input": {
        "why": "An assumption or input cell that nothing depends on is dead weight — it may mislead "
               "users into thinking a driver affects the model when it actually does not. It can also "
               "indicate a broken link where a formula was accidentally overwritten.",
        "cause": "Commonly occurs when an input was defined during model construction but the formulas "
                 "that referenced it were later deleted or rewritten to use a different cell.",
        "fix": "Either wire the input into the appropriate formula chain so it flows through the model, "
               "or remove it from the assumptions sheet to avoid confusion."
    },
    "Dangling Output": {
        "why": "A key output metric (IRR, NPV, MOIC, etc.) that contains a hard-coded value instead of "
               "a formula means the reported result does not update when assumptions change. This can "
               "produce silently stale or misleading outputs.",
        "cause": "Often caused by pasting values over formulas during a model hand-off, or by an analyst "
                 "manually typing a number to 'lock in' a result for a presentation.",
        "fix": "Replace the hard-coded value with the appropriate formula that derives the metric from "
               "model inputs. If the value is intentionally fixed, add a comment explaining why."
    }
}


def get_explanation(issue_type, error_value=None):
    """Returns the educational explanation for an issue type."""
    if issue_type not in ISSUE_EXPLANATIONS:
        return {"why": "", "cause": "", "fix": ""}

    explanation = ISSUE_EXPLANATIONS[issue_type].copy()

    # Handle specific error subtypes for Calculation Errors
    if issue_type == "Calculation Error" and error_value:
        causes = explanation["cause"]
        if isinstance(causes, dict):
            for err_code, err_cause in causes.items():
                if err_code in str(error_value):
                    explanation["cause"] = err_cause
                    break
            else:
                explanation["cause"] = "Unknown error type."

    return explanation


def _is_numeric(val):
    """Return True if val looks like a number (int, float, or numeric string)."""
    if isinstance(val, (int, float)):
        return True
    if not isinstance(val, str):
        return False
    try:
        float(val.replace(",", "").replace("$", "").replace("%", "").strip())
        return True
    except (ValueError, AttributeError):
        return False


# Sheet-level unit hints. Matched against top-N rows of each sheet so numeric
# values can be scaled up to raw USD before EDGAR comparison.
_UNIT_HINT_BILLIONS = re.compile(
    r"(?:\(|\b|^)\s*(?:\$\s*)?(?:in\s+)?(?:us(?:d)?\s+)?billions?\b"
    r"|\$?\s*bn\b|\$?\s*bns\b",
    re.IGNORECASE,
)
_UNIT_HINT_MILLIONS = re.compile(
    r"(?:\(|\b|^)\s*(?:\$\s*)?(?:in\s+)?(?:us(?:d)?\s+)?millions?\b"
    r"|\$\s*mm\b|\$\s*mn\b|\busd\s*mm\b|\busd\s*mn\b|\$mm\b",
    re.IGNORECASE,
)
_UNIT_HINT_THOUSANDS = re.compile(
    r"(?:\(|\b|^)\s*(?:\$\s*)?(?:in\s+)?(?:us(?:d)?\s+)?thousands?\b"
    r"|\$\s*000s?\b|\$000s?\b",
    re.IGNORECASE,
)


def _detect_sheet_scale(df, header_row_idx=None, max_probe=20):
    """Return a USD-unit multiplier (1_000, 1_000_000, or 1_000_000_000) based
    on explicit text hints in the top rows of the sheet. Returns None when no
    hint is found — callers should treat values as already in raw USD and let
    the validator's scale-mismatch heuristic catch off-by-1000 slips.

    Search spans ``max_probe`` rows from row 0 through ``header_row_idx`` (or
    ``max_probe``, whichever is larger).
    """
    if df is None or len(df) == 0:
        return None
    limit = max_probe
    if header_row_idx is not None:
        limit = max(limit, header_row_idx + 2)
    limit = min(limit, len(df))
    for r_idx in range(limit):
        row = df.iloc[r_idx].tolist()
        for v in row:
            if v is None or pd.isna(v):
                continue
            s = str(v)
            if _UNIT_HINT_BILLIONS.search(s):
                return 1_000_000_000.0
            if _UNIT_HINT_MILLIONS.search(s):
                return 1_000_000.0
            if _UNIT_HINT_THOUSANDS.search(s):
                return 1_000.0
    return None


def _detect_label_cols(row_values):
    """Count consecutive non-numeric, non-formula cells from column 0.

    Returns the number of leading label columns for this row.
    """
    count = 0
    for val in row_values:
        if pd.isna(val):
            # Empty cell in the label area is still part of labels
            count += 1
            continue
        s = str(val)
        if s.startswith("="):
            break
        if _is_numeric(val):
            break
        count += 1
    return count


class ModelAuditor:
    """
    The Analyst.

    Responsibility:
    1. Numerical Integrity: Checks if Balance Sheet balances.
    2. Logic Hygiene: Finds 'Plugs' (Hardcodes in formula rows).
    3. Structural Health: Flags Broken Links and External References.
    4. Circular References: Surfaces cycles detected in the dependency graph.
    5. Unreferenced Inputs: Finds dead assumptions with no dependents.
    6. Dangling Outputs: Flags key metrics that are hard-coded instead of formula-driven.
    """

    def __init__(self, ingestor, dependency_engine, hidden_sheets=None, audit_id=None,
                 ticker=None):
        self.ingestor = ingestor
        self.engine = dependency_engine
        self.graph = dependency_engine.graph if dependency_engine is not None else None
        self.hidden_sheets = hidden_sheets or set()
        self.audit_id = audit_id or ""
        self.ticker = ticker
        self.issues = []

    def run_all_checks(self):
        """Orchestrates the full suite of audit tests."""
        logger.info("[%s] Running hedge fund grade audit...", self.audit_id)

        self.check_external_links()
        self.detect_hardcoded_plugs()
        self.verify_balance_sheet_integrity()
        self._check_circular_refs()
        self._check_unreferenced_inputs()
        self._check_dangling_outputs()
        if self.ticker:
            self.check_edgar(self.ticker)

        logger.info("[%s] Audit complete. Found %d issues.", self.audit_id, len(self.issues))
        return self.issues

    def _add_issue(self, issue_type, severity, location, detail, error_value=None):
        """Helper to add an issue with its explanation."""
        explanation = get_explanation(issue_type, error_value)
        self.issues.append({
            "type": issue_type,
            "severity": severity,
            "location": location,
            "detail": detail,
            "why": explanation.get("why", ""),
            "cause": explanation.get("cause", ""),
            "fix": explanation.get("fix", "")
        })

    # ------------------------------------------------------------------
    # 1. External links & calculation errors
    # ------------------------------------------------------------------

    def check_external_links(self):
        """Scans for links to files outside this workbook (e.g., Bloomberg/FactSet)."""
        logger.info("Checking for external links & broken refs...")

        # 1. Check Graph for External Nodes
        if self.graph is not None:
            external_nodes = [n for n in self.graph.nodes if "EXT_LINK" in n]
            for node in external_nodes:
                self._add_issue(
                    issue_type="External Link",
                    severity="Medium",
                    location=node,
                    detail="Dependency on external workbook/source detected."
                )

        # 2. Check Values for Excel Errors (#REF!, #NAME?)
        for sheet_name, df in self.ingestor.sheets_values.items():
            if sheet_name in self.hidden_sheets:
                continue
            # Stack data to find errors efficiently
            errors = df.stack()
            # Filter for common Excel error strings
            error_cells = errors[errors.astype(str).str.contains(r"#REF!|#NAME\?|#VALUE!|#DIV/0!", regex=True)]

            for index, value in error_cells.items():
                row, col = index
                cell_ref = f"{sheet_name}!Row{row+1}:Col{col+1}"
                self._add_issue(
                    issue_type="Calculation Error",
                    severity="High",
                    location=cell_ref,
                    detail=f"Cell contains error value: {value}",
                    error_value=value
                )

    # ------------------------------------------------------------------
    # 2. Hard-coded plug detection
    # ------------------------------------------------------------------

    def detect_hardcoded_plugs(self):
        """
        Heuristic: If a row is mostly formulas but contains a hardcoded number
        in the projection columns, it's likely a 'Plug'.

        Label columns are auto-detected per row (consecutive non-numeric,
        non-formula cells from column 0). Threshold is 60%.  An additional
        secondary check flags any literal number whose immediate left AND right
        neighbors (same row) are formulas, regardless of the row threshold.
        """
        logger.info("Scanning for hard-coded plugs in projections...")

        for sheet_name, df_formulas in self.ingestor.sheets_formulas.items():
            if sheet_name in self.hidden_sheets:
                continue
            # Skip likely data dumps
            sheet_lower = sheet_name.lower()
            if any(kw in sheet_lower for kw in _DATA_SHEET_KEYWORDS):
                continue

            for idx, row in df_formulas.iterrows():
                full_row = row.tolist()

                # Auto-detect label columns for this row
                label_cols = _detect_label_cols(full_row)

                # Skip reconciliation / tie-out rows (check, balance, total,
                # subtotal, sum, audit, tieout, cross-check). A literal
                # between formulas on such a row is virtually always a
                # diagnostic value, not a plug.
                label_blob = " ".join(
                    str(c).lower().replace("-", "").replace(" ", "")
                    for c in full_row[:label_cols]
                    if pd.notna(c)
                )
                if any(kw in label_blob for kw in _PLUG_SKIP_KEYWORDS):
                    continue

                # Get the data portion after labels
                row_list = full_row[label_cols:]

                # --- Primary check: row-level threshold ---
                non_null = [(i, x) for i, x in enumerate(row_list) if pd.notna(x)]
                is_formula = [str(x).startswith('=') for _, x in non_null]

                if is_formula:
                    formula_count = sum(is_formula)
                    total_items = len(is_formula)

                    if total_items > 5 and (formula_count / total_items) > 0.60:
                        if formula_count < total_items:
                            hardcode_count = total_items - formula_count
                            plug_positions = [
                                i + label_cols + 1 for i, x in enumerate(row_list)
                                if pd.notna(x) and not str(x).startswith('=')
                            ]
                            self._add_issue(
                                issue_type="Hard-coded Plug",
                                severity="High",
                                location=f"{sheet_name}!Row{idx+1}",
                                detail=f"Row has {formula_count} formulas and {hardcode_count} "
                                       f"hardcodes in projection columns. Plug at col(s): {plug_positions}"
                            )
                            continue  # already flagged this row

                # --- Secondary check: sandwich detection ---
                # A literal number flanked by formulas on both sides
                for ci in range(1, len(full_row) - 1):
                    cell = full_row[ci]
                    if pd.isna(cell):
                        continue
                    if str(cell).startswith('='):
                        continue
                    if not _is_numeric(cell):
                        continue
                    left = full_row[ci - 1]
                    right = full_row[ci + 1]
                    if (pd.notna(left) and str(left).startswith('=') and
                            pd.notna(right) and str(right).startswith('=')):
                        self._add_issue(
                            issue_type="Hard-coded Plug",
                            severity="High",
                            location=f"{sheet_name}!Row{idx+1}:Col{ci+1}",
                            detail=f"Literal number sandwiched between formulas at column {ci+1}."
                        )

    # ------------------------------------------------------------------
    # 3. Balance sheet integrity
    # ------------------------------------------------------------------

    def verify_balance_sheet_integrity(self):
        """
        Locates Total Assets and Total Liabs+Equity and ensures variance is 0
        for each period individually.
        """
        logger.info("Verifying balance sheet logic...")

        bs_sheet = None
        for name in self.ingestor.sheets_values.keys():
            if _BS_PATTERNS.search(name):
                bs_sheet = name
                break

        if not bs_sheet:
            return  # No BS found

        df = self.ingestor.sheets_values[bs_sheet]

        # Locate the rows (Case insensitive search in first two columns)
        row_assets = None
        row_liabs_eq = None

        for col_idx in [0, 1]:
            for idx, val in df.iloc[:, col_idx].items():
                val_str = str(val)
                if row_assets is None and _ASSET_PATTERNS.search(val_str):
                    row_assets = idx
                if row_liabs_eq is None and _LIAB_EQ_PATTERNS.search(val_str):
                    row_liabs_eq = idx

        if row_assets is not None and row_liabs_eq is not None:
            # Extract the time series data (assumes data is to the right)
            assets_vals = pd.to_numeric(df.iloc[row_assets, 2:], errors='coerce').fillna(0)
            liabs_vals = pd.to_numeric(df.iloc[row_liabs_eq, 2:], errors='coerce').fillna(0)

            # Per-period variance
            variance = assets_vals - liabs_vals

            for period_idx, var_val in variance.items():
                abs_var = abs(var_val)
                # Proportional tolerance: 1bp of total assets or $1, whichever is larger
                asset_val = assets_vals.get(period_idx, 0)
                tolerance = max(1.0, abs(asset_val) * _BS_TOLERANCE_BPS)

                if abs_var > tolerance:
                    self._add_issue(
                        issue_type="Accounting Mismatch",
                        severity="Critical",
                        location=f"{bs_sheet}!Period {period_idx}",
                        detail=f"Balance Sheet does not balance in period {period_idx}. "
                               f"Variance: ${abs_var:,.2f} (tolerance: ${tolerance:,.2f})."
                    )

    # ------------------------------------------------------------------
    # 4. Circular reference reporting
    # ------------------------------------------------------------------

    def _check_circular_refs(self):
        """Surface cycles detected by the DependencyEngine."""
        if self.engine is None:
            return []

        logger.info("Checking for circular references...")

        import networkx as nx
        cycles = list(nx.simple_cycles(self.engine.graph))

        reported = 0
        for cycle in cycles:
            if reported >= 20:
                break
            cycle_path = " -> ".join(str(n) for n in cycle)
            self._add_issue(
                issue_type="Circular Reference",
                severity="Critical",
                location=cycle[0] if cycle else "Unknown",
                detail=f"Cycle of length {len(cycle)}: {cycle_path}"
            )
            reported += 1

        return self.issues[-reported:] if reported else []

    # ------------------------------------------------------------------
    # 5. Unreferenced inputs
    # ------------------------------------------------------------------

    def _check_unreferenced_inputs(self):
        """Find cells on assumptions/inputs/drivers sheets with zero outgoing edges."""
        if self.engine is None:
            return []

        logger.info("Checking for unreferenced inputs...")

        input_sheet_patterns = {"assumptions", "inputs", "drivers"}
        input_sheets = [
            name for name in self.ingestor.sheets_formulas.keys()
            if name.lower().strip() in input_sheet_patterns
        ]

        if not input_sheets:
            return []

        findings = []
        for sheet_name in input_sheets:
            df = self.ingestor.sheets_formulas[sheet_name]
            for row_idx, row in df.iterrows():
                for col_idx, cell_value in enumerate(row):
                    if pd.isna(cell_value):
                        continue
                    # We care about value cells (inputs), not formulas
                    if isinstance(cell_value, str) and cell_value.startswith("="):
                        continue
                    # Build the node id the same way DependencyEngine does
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(col_idx + 1)
                    node_id = f"{sheet_name}!{col_letter}{row_idx + 1}"

                    # Check if anything depends on this cell (out_degree in the graph)
                    if node_id in self.engine.graph:
                        if self.engine.graph.out_degree(node_id) > 0:
                            continue  # referenced — skip
                    # Node either absent from graph or has zero outgoing edges
                    self._add_issue(
                        issue_type="Unreferenced Input",
                        severity="Medium",
                        location=node_id,
                        detail="Input cell has no dependents in the model."
                    )
                    findings.append(node_id)

        return findings

    # ------------------------------------------------------------------
    # 6. Dangling outputs
    # ------------------------------------------------------------------

    def _check_dangling_outputs(self):
        """Flag key output metrics that are hard-coded instead of formula-driven."""
        if self.engine is None:
            return []

        logger.info("Checking for dangling outputs...")

        findings = []
        for sheet_name in self.ingestor.sheets_values.keys():
            df_vals = self.ingestor.sheets_values[sheet_name]
            df_forms = self.ingestor.sheets_formulas.get(sheet_name)
            if df_forms is None:
                continue

            for row_idx, row in df_vals.iterrows():
                # Scan label columns (first two) for output-metric keywords
                for label_col in [0, 1]:
                    if label_col >= len(row):
                        continue
                    label = str(row.iloc[label_col])
                    if not _OUTPUT_LABELS.search(label):
                        continue

                    # Check data cells to the right of label columns
                    for col_idx in range(2, len(row)):
                        cell_val = row.iloc[col_idx]
                        if pd.isna(cell_val):
                            continue
                        # Check the formula layer
                        form_val = df_forms.iloc[row_idx, col_idx] if col_idx < df_forms.shape[1] else None
                        if pd.isna(form_val):
                            continue
                        if isinstance(form_val, str) and form_val.startswith("="):
                            continue  # formula — OK
                        # Hard-coded value next to an output label
                        if _is_numeric(cell_val):
                            from openpyxl.utils import get_column_letter
                            col_letter = get_column_letter(col_idx + 1)
                            loc = f"{sheet_name}!{col_letter}{row_idx + 1}"
                            self._add_issue(
                                issue_type="Dangling Output",
                                severity="Medium",
                                location=loc,
                                detail=f"Output metric '{label.strip()}' contains hard-coded value "
                                       f"{cell_val} instead of a formula."
                            )
                            findings.append(loc)

        return findings

    # ------------------------------------------------------------------
    # 7. EDGAR historical-value check (optional; requires ticker)
    # ------------------------------------------------------------------

    def _classify_period_columns(self, df, header_row_idx=None):
        """Return {col_idx: (tag, year)} for each column whose header parses
        as a year. tag is 'historical', 'forecast', or 'unknown'.

        If ``header_row_idx`` is None, scans the first 10 rows for the one
        with the most year-parseable cells (requires >=3 to accept). Sellside
        models frequently put period labels below a title / unit banner row.
        """
        result = {}
        if df is None or len(df) == 0:
            return result
        if header_row_idx is None:
            header_row_idx = self._find_period_header_row(df)
            if header_row_idx is None:
                return result
        header = df.iloc[header_row_idx].tolist()
        current_year = pd.Timestamp.now().year
        for col_idx, cell in enumerate(header):
            if pd.isna(cell):
                continue
            text = str(cell)
            # Skip quarterly, half-year, and month-stamped columns — EDGAR
            # Core 6 values are annual (fp=FY) and comparisons would be off
            # by a factor of 4 (or more) if a quarterly column sneaks in.
            if _QUARTER_RE.search(text):
                continue
            m = _PERIOD_RE.search(text)
            if not m:
                continue
            if m.group(1):
                year = int(m.group(1))
                suffix = (m.group(2) or "").upper()
            else:
                # 2-digit FYXX short form: interpret as 2000-2099
                year = 2000 + int(m.group(3))
                suffix = (m.group(4) or "").upper()
            if suffix == "A":
                tag = "historical"
            elif suffix == "E":
                tag = "forecast"
            elif year < current_year:
                tag = "historical"
            elif year > current_year:
                tag = "forecast"
            else:
                tag = "unknown"
            result[col_idx] = (tag, year)
        return result

    @staticmethod
    def _find_period_header_row(df, max_probe=10, min_hits=3):
        """Return the index of the row with the most year-like cells, or None
        if no row has at least ``min_hits`` matches.
        """
        if df is None or len(df) == 0:
            return None
        best_idx = None
        best_count = 0
        for probe in range(min(max_probe, len(df))):
            row = df.iloc[probe].tolist()
            count = 0
            for v in row:
                if v is None or pd.isna(v):
                    continue
                if _PERIOD_RE.search(str(v)):
                    count += 1
            if count > best_count:
                best_count = count
                best_idx = probe
                if count >= 8:  # strong signal, no need to keep probing
                    break
        if best_count < min_hits:
            return None
        return best_idx

    def _collect_historical_samples(self):
        """Scan every sheet for rows whose label matches a Core 6 XBRL concept
        and whose column headers parse as historical periods.

        Returns a list of HistoricalCell objects (imported lazily to keep the
        auditor importable without the edgar module at module load).

        Sheet-level unit hints ("$ millions", "in thousands") are detected
        from the top rows and applied as a multiplier on numeric values so
        EDGAR comparisons can be made in raw USD. Cells carrying share counts
        or per-share values (EPS) are exempted from the sheet-level scale
        because their unit bucket is not USD.
        """
        from openpyxl.utils import get_column_letter

        from src.edgar_validator import HistoricalCell, _bucket_for, match_concepts

        samples = []
        for sheet_name, df in self.ingestor.sheets_values.items():
            if sheet_name in self.hidden_sheets:
                continue
            if df is None or len(df) < 2:
                continue
            header_row_idx = self._find_period_header_row(df)
            if header_row_idx is None:
                continue
            period_cols = self._classify_period_columns(df, header_row_idx=header_row_idx)
            if not period_cols:
                continue
            historical_cols = {
                cidx: year for cidx, (tag, year) in period_cols.items()
                if tag == "historical"
            }
            if not historical_cols:
                continue
            sheet_scale = _detect_sheet_scale(df, header_row_idx=header_row_idx)
            # Start scanning rows below the detected header
            for row_idx in range(header_row_idx + 1, len(df)):
                row = df.iloc[row_idx]
                label_col_count = _detect_label_cols(row.tolist())
                if label_col_count == 0:
                    continue
                # Walk label cells right-to-left; use the first non-empty one
                # as the canonical label. Sellside models often leave trailing
                # label columns blank (e.g. [label, None, numbers...]), and
                # hierarchical layouts push the most specific label rightmost.
                label_str = None
                for li in range(label_col_count - 1, -1, -1):
                    cand = row.iloc[li]
                    if pd.isna(cand):
                        continue
                    text = str(cand).strip()
                    if not text:
                        continue
                    label_str = text
                    break
                if label_str is None:
                    continue
                concepts = match_concepts(label_str)
                if not concepts:
                    continue
                for col_idx, year in historical_cols.items():
                    if col_idx >= len(row):
                        continue
                    val = row.iloc[col_idx]
                    if pd.isna(val) or not _is_numeric(val):
                        continue
                    try:
                        numeric_val = float(
                            str(val).replace(",", "").replace("$", "").replace("%", "").strip()
                        )
                    except (ValueError, AttributeError):
                        continue
                    # Apply sheet-level unit scale for USD-bucket concepts only.
                    # Per-share and share-count values (EPS, diluted shares) are
                    # already in their own bucket and are not re-scaled.
                    if sheet_scale and sheet_scale != 1.0:
                        primary_bucket = _bucket_for(concepts[0])
                        if primary_bucket == "USD":
                            numeric_val *= sheet_scale
                    samples.append(HistoricalCell(
                        sheet=sheet_name,
                        row_label=label_str,
                        col_letter=get_column_letter(col_idx + 1),
                        row_idx=row_idx,
                        col_idx=col_idx,
                        year=year,
                        value=numeric_val,
                        concepts=concepts,
                    ))
        return samples

    def check_edgar(self, ticker, validator=None):
        """Validate historical cells against SEC EDGAR values.

        When ticker is falsy, does nothing. On any resolution / network
        failure, appends a single 'EDGAR Check Skipped' issue rather than
        raising.
        """
        if not ticker or not str(ticker).strip():
            return

        logger.info("Running EDGAR check for ticker %s...", ticker)

        if validator is None:
            from src.edgar_validator import EDGARValidator
            validator = EDGARValidator()

        samples = self._collect_historical_samples()
        new_issues = validator.validate(ticker, samples)
        self.issues.extend(new_issues)
