import logging
import re

import networkx as nx
import pandas as pd
from openpyxl.formula.tokenizer import Token, Tokenizer
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string

logger = logging.getLogger(__name__)

# Regex to detect a cell reference (with optional $ anchors and optional sheet prefix).
# Matches patterns like A1, $A$1, A1:B5, Sheet1!A1, 'Sheet 1'!$A$1:$B$5
_CELL_RE = re.compile(
    r"^\$?[A-Za-z]{1,3}\$?\d+(?::\$?[A-Za-z]{1,3}\$?\d+)?$"
)

# Strict cell-reference pattern for distinguishing bare refs from named ranges.
# Matches A1, $B$2, AA100, Q1, etc. but NOT Rate2024 or Revenue.
_CELL_REF_PATTERN = re.compile(r"^\$?[A-Z]{1,3}\$?\d+$", re.IGNORECASE)

# Max cells expanded from a single range to prevent OOM on whole-column refs
_MAX_RANGE_CELLS = 10_000

# Max cycles stored in circular_reference_cells
_MAX_CYCLES = 50


class DependencyEngine:
    """
    The Logic Map.

    Responsibility:
    1. Parse Excel formulas to understand "Precedents" (Inputs) and "Dependents" (Outputs).
    2. Build a NetworkX Directed Graph where:
       - Nodes = Cells (e.g., 'Summary!B10')
       - Edges = Data Flow (Source -> Target)
    3. Detect Structural Risks (Circular References, Broken Chains).
    """

    def __init__(self, sheets_formulas: dict, audit_id: str = None):
        self.raw_formulas = sheets_formulas
        self.audit_id = audit_id or ""
        self.graph = nx.DiGraph()
        self.node_count = 0
        self._defined_names = {}  # name -> "Sheet!Range" mapping
        self._unresolved_named_ranges = 0  # count of names that failed to parse

    def set_defined_names(self, defined_names):
        """Accept openpyxl workbook.defined_names for resolution during parsing.

        Parameters
        ----------
        defined_names : openpyxl.workbook.defined_name.DefinedNameList
            The workbook's defined_names object.  Each entry is resolved to
            its first destination as ``"Sheet!CellOrRange"`` and stored for
            lookup during formula tokenisation.
        """
        self._defined_names = {}
        for dn in defined_names.definedName:
            try:
                # Skip external / reserved names
                if dn.is_external or dn.is_reserved:
                    continue
                for sheet_title, coord in dn.destinations:
                    # Store first destination only; strip $ signs for
                    # consistency with node IDs elsewhere in the graph.
                    clean_coord = coord.replace("$", "")
                    self._defined_names[dn.name.lower()] = f"{sheet_title}!{clean_coord}"
                    break
            except (ValueError, KeyError, AttributeError):
                # Some names (e.g. formula-based) have no simple destination
                self._unresolved_named_ranges += 1

    def build_graph(self):
        """Iterates through every cell in every sheet to map dependencies."""
        logger.info("[%s] Building dependency graph...", self.audit_id)

        for sheet_name, df in self.raw_formulas.items():
            # Iterate through the DataFrame (0-indexed)
            for row_idx, row in df.iterrows():
                for col_idx, cell_value in enumerate(row):

                    # We only care about Formulas (strings starting with =)
                    # Also handle array formulas stored as {=...}
                    if isinstance(cell_value, str) and (
                        cell_value.startswith("=") or cell_value.startswith("{=")
                    ):
                        target_node = self._get_node_id(sheet_name, row_idx, col_idx)
                        self._parse_formula(target_node, cell_value, sheet_name)
                        self.node_count += 1

        logger.info("Graph built. Total calculation nodes: %d", self.node_count)
        logger.info("Total dependencies mapped: %d", self.graph.number_of_edges())

    def _get_node_id(self, sheet, row_idx, col_idx):
        """Standardizes Node IDs: 'SheetName!A1'"""
        col_letter = get_column_letter(col_idx + 1)
        row_num = row_idx + 1
        return f"{sheet}!{col_letter}{row_num}"

    def _parse_formula(self, target_node, formula_str, current_sheet):
        """
        Uses OpenPyXL Tokenizer to find all precedents in a formula.
        Example: =SUM(PFNA!A1:A5) -> Creates edges from PFNA!A1...A5 to Target.

        Handles:
        - Range expansion (A1:C3 -> individual cells)
        - Array formulas ({=...} braces stripped before tokenising)
        - Named range resolution (via set_defined_names)
        - Table references (Table1[Column] stored as table_ref nodes)
        """
        try:
            # --- Fix 3: Array formula handling ---
            # Strip leading { and trailing } so the tokenizer parses normally
            cleaned = formula_str
            if cleaned.startswith("{") and cleaned.endswith("}"):
                cleaned = cleaned[1:-1]

            tok = Tokenizer(cleaned)

            for t in tok.items:
                if t.type == Token.OPERAND and t.subtype == Token.RANGE:
                    ref_value = t.value

                    # --- Fix 4: Table reference handling ---
                    if "[" in ref_value and "]" in ref_value and "!" not in ref_value:
                        # Table reference like Table1[Column] or Table1[#Headers]
                        table_node = f"TABLE_REF:{ref_value}"
                        self.graph.add_node(table_node, type="table_ref")
                        self.graph.add_edge(table_node, target_node)
                        continue

                    # --- Fix 2: Named range resolution ---
                    # Strip any sheet qualifier first for the check
                    bare_ref = ref_value.split("!")[-1] if "!" in ref_value else ref_value

                    # Skip Excel built-in names (e.g. _xlnm.Print_Area)
                    if bare_ref.startswith("_xlnm"):
                        continue

                    # Use cell-ref pattern to distinguish real cell refs (Q1, AA100)
                    # from named ranges (Rate2024, Revenue).
                    if not _CELL_REF_PATTERN.match(bare_ref) and bare_ref.lower() in self._defined_names:
                        ref_value = self._defined_names[bare_ref.lower()]

                    # Delegate to _add_edge which now handles range expansion
                    self._add_edge(target_node, ref_value, current_sheet)

        except Exception:
            # If tokenizer fails, we flag the node but don't crash
            self.graph.add_node(target_node, status="parse_error")

    @staticmethod
    def _expand_range(cell_range):
        """Expand a range string like 'A1:C3' into individual cell strings.

        Returns a list of cell strings. If the range would exceed
        _MAX_RANGE_CELLS, returns the original range string wrapped in a
        single-element list (falls back to block-level node).

        Single cell references (no colon) are returned as-is.
        """
        # Strip $ anchors for uniform handling
        clean = cell_range.replace("$", "")

        if ":" not in clean:
            return [clean]

        start_str, end_str = clean.split(":", 1)

        try:
            start_col_str, start_row = coordinate_from_string(start_str)
            end_col_str, end_row = coordinate_from_string(end_str)
        except (ValueError, TypeError):
            # Unparseable (e.g. whole-column A:A) -- keep as block
            return [cell_range.replace("$", "")]

        start_col = column_index_from_string(start_col_str)
        end_col = column_index_from_string(end_col_str)

        # Ensure ordering
        if start_col > end_col:
            start_col, end_col = end_col, start_col
        if start_row > end_row:
            start_row, end_row = end_row, start_row

        num_cells = (end_row - start_row + 1) * (end_col - start_col + 1)
        if num_cells > _MAX_RANGE_CELLS:
            return [cell_range.replace("$", "")]

        cells = []
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                cells.append(f"{get_column_letter(c)}{r}")
        return cells

    def _add_edge(self, target_node, source_ref, current_sheet):
        """
        Resolves the source reference (handling implied sheet names) and adds to Graph.
        Ranges are expanded into individual cell nodes.
        """
        # 1. Handle External Links (Don't map, just tag)
        if "[" in source_ref and "]" in source_ref:
            self.graph.add_edge(f"EXT_LINK:{source_ref}", target_node)
            return

        # 2. Parse Sheet vs Cell
        if "!" in source_ref:
            source_sheet, source_range = source_ref.split("!", 1)
            source_sheet = source_sheet.replace("'", "")  # Clean quotes
        else:
            source_sheet = current_sheet
            source_range = source_ref

        # 3. Expand range into individual cells and add edges
        expanded = self._expand_range(source_range)
        for cell_ref in expanded:
            source_node = f"{source_sheet}!{cell_ref}"
            self.graph.add_edge(source_node, target_node)

    def analyze_structure(self):
        """Returns high-level structural risks.

        The returned dict includes:
        - circular_references: int count of cycles
        - circular_reference_cells: list of lists, each inner list contains
          the cell addresses in one cycle (capped at first 50 cycles)
        - orphaned_calculations: list of node IDs with outgoing edges but no
          downstream consumers
        - complexity_score: edges / (nodes + 1)
        - unresolved_table_refs: list of TABLE_REF node IDs that cannot be
          resolved to actual cells
        """
        cycles = list(nx.simple_cycles(self.graph))
        table_refs = [n for n in self.graph.nodes if n.startswith("TABLE_REF:")]
        return {
            "circular_references": len(cycles),
            "circular_reference_cells": [list(c) for c in cycles[:_MAX_CYCLES]],
            "orphaned_calculations": [n for n, d in self.graph.out_degree() if d == 0 and self.graph.in_degree(n) > 0],
            "complexity_score": self.graph.number_of_edges() / (self.graph.number_of_nodes() + 1),
            "unresolved_table_refs": table_refs,
        }

# Unit Test
if __name__ == "__main__":
    # Mock data to test logic without loading full PEP model
    mock_data = {
        "Summary": pd.DataFrame([["=PFNA!A1 + 100", 50]]),
        "PFNA": pd.DataFrame([[100, 200]])
    }

    eng = DependencyEngine(mock_data)
    eng.build_graph()
    stats = eng.analyze_structure()
    logger.info("Cycles detected: %d", stats['circular_references'])
    logger.info("Cycle cells: %s", stats['circular_reference_cells'])
