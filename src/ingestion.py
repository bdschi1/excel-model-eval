import logging
import os
import threading
import warnings
import zipfile

import openpyxl
import openpyxl.utils.exceptions
import pandas as pd

logger = logging.getLogger(__name__)

# Suppress warnings for "extension is not supported" often found in complex financial models
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


class IngestionTimeout(Exception):
    pass


class ModelIngestor:
    """
    The Gateway for the Excel Audit Tool.

    Responsibility:
    1. Load the target workbook in two modes: Values (Math) and Formulas (Logic).
    2. Convert every tab into a high-fidelity DataFrame.
    3. Handle 'External Link' errors gracefully (e.g., FactSet/Bloomberg links).
    """

    MAX_FILE_SIZE_MB = 50
    MAX_ROWS = 100_000
    MAX_COLS = 5_000

    def __init__(self, file_path: str, audit_id: str = None):
        self.file_path = file_path
        self.filename = os.path.basename(file_path)
        self.audit_id = audit_id or ""

        # Dual-State Storage
        self.wb_values = None     # Holds calculated numbers (e.g., 150.5)
        self.wb_formulas = None   # Holds logic strings (e.g., =SUM(A1:A5))

        # Dictionary Store: { "SheetName": pd.DataFrame }
        self.sheets_values = {}
        self.sheets_formulas = {}

        # Hidden sheet tracking
        self.hidden_sheets = set()

        # Defined names (named ranges) from the workbook
        self.defined_names = None

        # Error Tracking
        self.load_errors = []

    def _load_with_timeout(self, load_func, timeout=120):
        """Wraps a callable with a threading-based timeout for Streamlit compatibility."""
        result = [None]
        error = [None]

        def target():
            try:
                result[0] = load_func()
            except (
                OSError,
                ValueError,
                KeyError,
                zipfile.BadZipFile,
                openpyxl.utils.exceptions.InvalidFileException,
            ) as e:
                error[0] = e

        t = threading.Thread(target=target)
        t.daemon = True
        t.start()
        t.join(timeout)
        if t.is_alive():
            # Release workbook reference to reduce memory impact of orphaned thread
            if result[0] is not None:
                del result[0]
            logger.warning(
                "Ingestion thread still alive after %ds timeout; "
                "daemon thread will be abandoned", timeout
            )
            raise IngestionTimeout(f"Ingestion timed out after {timeout}s")
        if error[0]:
            raise error[0]
        return result[0]

    def _validate_magic_bytes(self) -> bool:
        """Validate file magic bytes before loading. Returns True if valid."""
        ext = os.path.splitext(self.file_path)[1].lower()
        if ext == '.csv':
            return True

        try:
            with open(self.file_path, 'rb') as f:
                header = f.read(8)
        except OSError as e:
            self.load_errors.append(f"Cannot read file header: {e}")
            return False

        if ext in ('.xlsx', '.xlsm'):
            # ZIP format: starts with PK (0x504B)
            if len(header) < 2 or header[:2] != b'PK':
                self.load_errors.append(
                    f"Invalid file format: {self.filename} does not have valid ZIP/OOXML header"
                )
                return False
        elif ext == '.xls':
            # OLE2 format: starts with 0xD0CF11E0
            if len(header) < 4 or header[:4] != b'\xd0\xcf\x11\xe0':
                self.load_errors.append(
                    f"Invalid file format: {self.filename} does not have valid OLE2 header"
                )
                return False

        return True

    def ingest(self) -> bool:
        """
        Orchestrates the full loading process.
        Returns True if the critical mass of the model is loaded.
        """
        logger.info("[%s] Starting ingestion for: %s", self.audit_id, self.filename)

        if not os.path.exists(self.file_path):
            logger.error("File not found: %s", self.file_path)
            return False

        # File size check
        file_size_mb = os.path.getsize(self.file_path) / (1024 * 1024)
        if file_size_mb > self.MAX_FILE_SIZE_MB:
            error_msg = (
                f"File too large: {file_size_mb:.1f} MB exceeds "
                f"{self.MAX_FILE_SIZE_MB} MB limit"
            )
            self.load_errors.append(error_msg)
            logger.error("%s", error_msg)
            return False

        # Magic bytes validation
        if not self._validate_magic_bytes():
            logger.error("File format validation failed for: %s", self.filename)
            return False

        try:
            # 1. Load Values (Numerical Layer)
            logger.info("Loading numerical layer (values)...")
            self.wb_values = self._load_with_timeout(
                lambda: openpyxl.load_workbook(
                    self.file_path, data_only=True, read_only=False, keep_vba=False
                )
            )

            # Store defined names (named ranges) for downstream use
            self.defined_names = self.wb_values.defined_names

            # 2. Load Logic (Formula Layer)
            logger.info("Loading logic layer (formulas)...")
            self.wb_formulas = self._load_with_timeout(
                lambda: openpyxl.load_workbook(
                    self.file_path, data_only=False, read_only=False, keep_vba=False
                )
            )

            # 3. Track hidden sheets
            for name in self.wb_values.sheetnames:
                if self.wb_values[name].sheet_state != 'visible':
                    self.hidden_sheets.add(name)

            # 4. Process Sheets
            self._process_all_sheets()

            return True

        except IngestionTimeout as e:
            error_msg = f"INGESTION TIMEOUT: {str(e)}"
            self.load_errors.append(error_msg)
            logger.error("[%s] %s", self.audit_id, error_msg)
            return False
        except openpyxl.utils.exceptions.InvalidFileException as e:
            error_msg = f"Invalid Excel file: {str(e)}"
            self.load_errors.append(error_msg)
            logger.error("%s", error_msg)
            return False
        except zipfile.BadZipFile as e:
            error_msg = f"Corrupt ZIP/Excel file: {str(e)}"
            self.load_errors.append(error_msg)
            logger.error("%s", error_msg)
            return False
        except KeyError as e:
            error_msg = f"Missing expected key during ingestion: {str(e)}"
            self.load_errors.append(error_msg)
            logger.error("%s", error_msg)
            return False
        except ValueError as e:
            error_msg = f"Value error during ingestion: {str(e)}"
            self.load_errors.append(error_msg)
            logger.error("%s", error_msg)
            return False
        except OSError as e:
            error_msg = f"OS/IO error during ingestion: {str(e)}"
            self.load_errors.append(error_msg)
            logger.error("%s", error_msg)
            return False
        except Exception as e:
            error_msg = f"CRITICAL INGESTION FAILURE ({type(e).__name__}): {str(e)}"
            self.load_errors.append(error_msg)
            logger.error("[%s] %s", self.audit_id, error_msg)
            return False

    def _process_all_sheets(self):
        """Iterates through all sheets and converts them to DataFrames."""
        # We assume sheet names are identical in both load modes
        sheet_names = self.wb_values.sheetnames
        logger.info("Processing %d sheets...", len(sheet_names))

        for sheet in sheet_names:
            try:
                # Extract Values
                df_val = self._sheet_to_df(self.wb_values[sheet], sheet)
                self.sheets_values[sheet] = df_val

                # Extract Formulas
                df_form = self._sheet_to_df(self.wb_formulas[sheet], sheet)
                self.sheets_formulas[sheet] = df_form

            except KeyError as e:
                error_msg = f"Sheet '{sheet}' not found in one workbook mode: {str(e)}"
                self.load_errors.append(error_msg)
                logger.warning("%s", error_msg)
            except ValueError as e:
                error_msg = f"Value error parsing sheet '{sheet}': {str(e)}"
                self.load_errors.append(error_msg)
                logger.warning("%s", error_msg)
            except Exception as e:
                # Last resort — log error but do not crash the program
                error_msg = f"Failed to parse sheet '{sheet}': {str(e)}"
                self.load_errors.append(error_msg)
                logger.warning("%s", error_msg)

    def _sheet_to_df(self, worksheet, sheet_name: str = "") -> pd.DataFrame:
        """
        Converts an openpyxl worksheet to a Pandas DataFrame.
        Does NOT use the first row as header automatically to preserve structure.
        Caps rows at MAX_ROWS and columns at MAX_COLS.
        """
        data = []
        truncated_rows = False
        truncated_cols = False

        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True)):
            if row_idx >= self.MAX_ROWS:
                truncated_rows = True
                break
            if len(row) > self.MAX_COLS:
                data.append(row[:self.MAX_COLS])
                truncated_cols = True
            else:
                data.append(row)

        if truncated_rows:
            total_rows = worksheet.max_row or 0
            msg = (
                f"Sheet '{sheet_name}': truncated to "
                f"{self.MAX_ROWS:,} rows (sheet has {total_rows:,})"
            )
            self.load_errors.append(msg)
            logger.warning("%s", msg)
        if truncated_cols:
            total_cols = worksheet.max_column or 0
            msg = (
                f"Sheet '{sheet_name}': truncated to "
                f"{self.MAX_COLS:,} columns (sheet has {total_cols:,})"
            )
            self.load_errors.append(msg)
            logger.warning("%s", msg)

        if not data:
            return pd.DataFrame()

        # Create DataFrame with 0-based integer index/columns (A=0, B=1...)
        # This aligns perfectly with the Dependency Graph's (row, col) coordinate system.
        df = pd.DataFrame(data)
        return df

    def get_ingestion_report(self):
        """Returns a summary of what was loaded and any errors encountered."""
        return {
            "total_sheets": len(self.sheets_values),
            "sheet_names": list(self.sheets_values.keys()),
            "hidden_sheets": list(self.hidden_sheets),
            "errors": self.load_errors,
            "status": "Success" if not self.load_errors else "Partial Success"
        }


# For testing this module in isolation
if __name__ == "__main__":
    # Test on the local file assuming standard dir structure
    test_path = os.path.join(os.getcwd(), "data", "pep model for ai eval test.xlsx")
    ingestor = ModelIngestor(test_path)
    if ingestor.ingest():
        report = ingestor.get_ingestion_report()
        logger.info("Ingestion complete. Loaded sheets: %d", report['total_sheets'])
        if report['errors']:
            logger.warning("Errors: %s", report['errors'])
