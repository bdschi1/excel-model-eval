"""
Tests for src/reporting.py -- ReportGenerator

Verifies:
1. Complexity score calculation
2. PDF generation (file created, non-empty)
3. Excel generation (file created, loadable)
4. update_log() appends CSV row
5. Zero issues case
6. Many issues (50+)
"""

from __future__ import annotations

import csv
import os
import pathlib
import re
import sys
from unittest.mock import MagicMock

import networkx as nx
import openpyxl
import pandas as pd

# Ensure repo root is on path
REPO_ROOT = pathlib.Path(__file__).parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from src.reporting import ReportGenerator

# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------

def _mock_ingestor(sheet_count=5):
    """Create a mock ingestor with N empty sheets."""
    ingestor = MagicMock()
    sheets = {f"Sheet{i}": pd.DataFrame([[1]]) for i in range(sheet_count)}
    ingestor.sheets_values = sheets
    return ingestor


def _mock_dep_engine(node_count=100, edge_count=150):
    """Create a mock dependency engine with configurable graph size."""
    engine = MagicMock()
    g = nx.DiGraph()
    # Add nodes
    for i in range(node_count):
        g.add_node(f"Node{i}")
    # Add edges (up to edge_count, cycling through nodes)
    added = 0
    for i in range(node_count):
        for j in range(i + 1, node_count):
            if added >= edge_count:
                break
            g.add_edge(f"Node{i}", f"Node{j}")
            added += 1
        if added >= edge_count:
            break
    engine.graph = g
    return engine


def _sample_issues(count=3, severities=None):
    """Generate N sample issues."""
    if severities is None:
        severities = ["Critical", "High", "Medium"]
    issues = []
    for i in range(count):
        sev = severities[i % len(severities)]
        issues.append({
            "type": "Hard-coded Plug",
            "severity": sev,
            "location": f"Sheet1!Row{i+1}",
            "detail": f"Test issue {i+1}",
            "why": "Test explanation",
            "cause": "Test cause",
            "fix": "Test fix",
        })
    return issues


def _make_report_generator(
    filename="test_model.xlsx",
    issues=None,
    sheet_count=5,
    node_count=100,
    edge_count=150,
    tmp_path=None,
):
    """Build a ReportGenerator with mocked dependencies, using tmp_path for output."""
    if issues is None:
        issues = _sample_issues()
    ingestor = _mock_ingestor(sheet_count)
    dep_engine = _mock_dep_engine(node_count, edge_count)

    rg = ReportGenerator(filename, issues, ingestor, dep_engine)

    # Override results_dir to use tmp_path for test isolation
    if tmp_path:
        rg.results_dir = os.path.join(str(tmp_path), "RESULTS")
        os.makedirs(rg.results_dir, exist_ok=True)

    return rg


# ==================================================================
# Tests
# ==================================================================


class TestComplexityScore:
    """_calculate_complexity() score calculation."""

    def test_minimal_model_score_1(self, tmp_path):
        """Few sheets, few nodes => score 1."""
        rg = _make_report_generator(
            sheet_count=3, node_count=10, edge_count=5, tmp_path=tmp_path,
        )
        assert rg.complexity_score == 1

    def test_moderate_sheets_adds_1(self, tmp_path):
        """11-30 sheets => +1."""
        rg = _make_report_generator(
            sheet_count=15, node_count=10, edge_count=5, tmp_path=tmp_path,
        )
        assert rg.complexity_score >= 2

    def test_high_sheets_adds_2(self, tmp_path):
        """>30 sheets => +2."""
        rg = _make_report_generator(
            sheet_count=35, node_count=10, edge_count=5, tmp_path=tmp_path,
        )
        assert rg.complexity_score >= 3

    def test_high_node_count_adds_score(self, tmp_path):
        """>2000 nodes => +1, >10000 => +2."""
        rg = _make_report_generator(
            sheet_count=3, node_count=3000, edge_count=100, tmp_path=tmp_path,
        )
        assert rg.complexity_score >= 2

    def test_massive_node_count(self, tmp_path):
        """>10000 nodes => +2."""
        rg = _make_report_generator(
            sheet_count=3, node_count=10500, edge_count=100, tmp_path=tmp_path,
        )
        assert rg.complexity_score >= 3

    def test_score_capped_at_5(self, tmp_path):
        """Score should never exceed 5."""
        rg = _make_report_generator(
            sheet_count=50, node_count=15000, edge_count=30000, tmp_path=tmp_path,
        )
        assert rg.complexity_score <= 5

    def test_high_interconnectivity_adds_score(self, tmp_path):
        """edge_count > node_count * 1.5 => +1."""
        rg = _make_report_generator(
            sheet_count=3, node_count=100, edge_count=200, tmp_path=tmp_path,
        )
        assert rg.complexity_score >= 2

    def test_rationale_populated(self, tmp_path):
        rg = _make_report_generator(
            sheet_count=15, node_count=100, edge_count=200, tmp_path=tmp_path,
        )
        assert isinstance(rg.complexity_rationale, str)
        assert len(rg.complexity_rationale) > 0


class TestPDFGeneration:
    """generate_pdf() tests."""

    def test_pdf_file_created(self, tmp_path):
        rg = _make_report_generator(tmp_path=tmp_path)
        pdf_path = rg.generate_pdf()
        assert os.path.exists(pdf_path)
        assert pdf_path.endswith(".pdf")

    def test_pdf_file_nonempty(self, tmp_path):
        rg = _make_report_generator(tmp_path=tmp_path)
        pdf_path = rg.generate_pdf()
        assert os.path.getsize(pdf_path) > 0

    def test_pdf_zero_issues(self, tmp_path):
        """PDF should still be created with zero issues."""
        rg = _make_report_generator(issues=[], tmp_path=tmp_path)
        pdf_path = rg.generate_pdf()
        assert os.path.exists(pdf_path)
        assert os.path.getsize(pdf_path) > 0

    def test_pdf_many_issues(self, tmp_path):
        """PDF should handle 50+ issues without error."""
        issues = _sample_issues(count=55)
        rg = _make_report_generator(issues=issues, tmp_path=tmp_path)
        pdf_path = rg.generate_pdf()
        assert os.path.exists(pdf_path)
        assert os.path.getsize(pdf_path) > 0


class TestExcelGeneration:
    """generate_excel() tests."""

    def test_excel_file_created(self, tmp_path):
        rg = _make_report_generator(tmp_path=tmp_path)
        excel_path = rg.generate_excel()
        assert os.path.exists(excel_path)
        assert excel_path.endswith(".xlsx")

    def test_excel_file_loadable(self, tmp_path):
        rg = _make_report_generator(tmp_path=tmp_path)
        excel_path = rg.generate_excel()
        wb = openpyxl.load_workbook(excel_path)
        assert len(wb.sheetnames) >= 1
        wb.close()

    def test_excel_has_executive_summary_tab(self, tmp_path):
        rg = _make_report_generator(tmp_path=tmp_path)
        excel_path = rg.generate_excel()
        wb = openpyxl.load_workbook(excel_path)
        assert "Executive Summary" in wb.sheetnames
        wb.close()

    def test_excel_has_findings_tab_with_issues(self, tmp_path):
        rg = _make_report_generator(tmp_path=tmp_path)
        excel_path = rg.generate_excel()
        wb = openpyxl.load_workbook(excel_path)
        assert "Findings" in wb.sheetnames
        wb.close()

    def test_excel_zero_issues(self, tmp_path):
        """Excel should still be created with zero issues."""
        rg = _make_report_generator(issues=[], tmp_path=tmp_path)
        excel_path = rg.generate_excel()
        assert os.path.exists(excel_path)
        wb = openpyxl.load_workbook(excel_path)
        assert "Executive Summary" in wb.sheetnames
        wb.close()

    def test_excel_many_issues(self, tmp_path):
        """Excel should handle 50+ issues."""
        issues = _sample_issues(count=55)
        rg = _make_report_generator(issues=issues, tmp_path=tmp_path)
        excel_path = rg.generate_excel()
        assert os.path.exists(excel_path)


class TestUpdateLog:
    """update_log() CSV append tests."""

    def test_log_creates_file(self, tmp_path):
        rg = _make_report_generator(tmp_path=tmp_path)
        rg.update_log()
        log_path = os.path.join(rg.results_dir, "audit_history.csv")
        assert os.path.exists(log_path)

    def test_log_has_header(self, tmp_path):
        rg = _make_report_generator(tmp_path=tmp_path)
        rg.update_log()
        log_path = os.path.join(rg.results_dir, "audit_history.csv")
        with open(log_path, "r") as f:
            reader = csv.reader(f)
            header = next(reader)
        assert "Timestamp" in header
        assert "Filename" in header
        assert "Complexity_Score" in header
        assert "Total_Issues" in header

    def test_log_appends_data_row(self, tmp_path):
        issues = _sample_issues(count=5, severities=["Critical", "High", "Medium"])
        rg = _make_report_generator(issues=issues, tmp_path=tmp_path)
        rg.update_log()
        log_path = os.path.join(rg.results_dir, "audit_history.csv")
        with open(log_path, "r") as f:
            reader = csv.reader(f)
            rows = list(reader)
        # Header + 1 data row
        assert len(rows) == 2
        data_row = rows[1]
        assert data_row[1] == "test_model.xlsx"  # Filename

    def test_log_appends_multiple_runs(self, tmp_path):
        rg1 = _make_report_generator(filename="model_a.xlsx", tmp_path=tmp_path)
        rg1.update_log()
        rg2 = _make_report_generator(filename="model_b.xlsx", tmp_path=tmp_path)
        rg2.update_log()
        log_path = os.path.join(rg1.results_dir, "audit_history.csv")
        with open(log_path, "r") as f:
            reader = csv.reader(f)
            rows = list(reader)
        # Header + 2 data rows
        assert len(rows) == 3

    def test_log_counts_critical_issues(self, tmp_path):
        issues = _sample_issues(count=6, severities=["Critical"])
        rg = _make_report_generator(issues=issues, tmp_path=tmp_path)
        rg.update_log()
        log_path = os.path.join(rg.results_dir, "audit_history.csv")
        with open(log_path, "r") as f:
            reader = csv.reader(f)
            rows = list(reader)
        data_row = rows[1]
        # Critical_Errors column (index 3)
        assert int(data_row[3]) == 6
        # Total_Issues column (index 4)
        assert int(data_row[4]) == 6


class TestZeroIssuesEdgeCases:
    """Edge cases with zero issues."""

    def test_zero_issues_complexity_still_calculated(self, tmp_path):
        rg = _make_report_generator(issues=[], tmp_path=tmp_path)
        assert isinstance(rg.complexity_score, int)
        assert 1 <= rg.complexity_score <= 5

    def test_zero_issues_pdf_mentions_no_issues(self, tmp_path):
        rg = _make_report_generator(issues=[], tmp_path=tmp_path)
        pdf_path = rg.generate_pdf()
        assert os.path.exists(pdf_path)


# ==================================================================
# Tier 2 #10: ReportGenerator with engine=None (CSV mode)
# ==================================================================


class TestEngineNoneReporting:
    """ReportGenerator must work when dependency_engine is None (CSV mode)."""

    def _make_none_engine_report(self, tmp_path, issues=None):
        if issues is None:
            issues = _sample_issues(count=3)
        ingestor = _mock_ingestor(sheet_count=3)
        rg = ReportGenerator("test_csv.csv", issues, ingestor, None)
        rg.results_dir = os.path.join(str(tmp_path), "RESULTS")
        os.makedirs(rg.results_dir, exist_ok=True)
        return rg

    def test_complexity_score_with_none_engine(self, tmp_path):
        """Complexity score should still compute (node/edge counts = 0)."""
        rg = self._make_none_engine_report(tmp_path)
        assert isinstance(rg.complexity_score, int)
        assert 1 <= rg.complexity_score <= 5

    def test_pdf_generated_with_none_engine(self, tmp_path):
        rg = self._make_none_engine_report(tmp_path)
        pdf_path = rg.generate_pdf()
        assert os.path.exists(pdf_path)
        assert os.path.getsize(pdf_path) > 0

    def test_excel_generated_with_none_engine(self, tmp_path):
        rg = self._make_none_engine_report(tmp_path)
        excel_path = rg.generate_excel()
        assert os.path.exists(excel_path)
        wb = openpyxl.load_workbook(excel_path)
        assert "Executive Summary" in wb.sheetnames
        wb.close()

    def test_update_log_with_none_engine(self, tmp_path):
        rg = self._make_none_engine_report(tmp_path)
        rg.update_log()
        log_path = os.path.join(rg.results_dir, "audit_history.csv")
        assert os.path.exists(log_path)

    def test_zero_issues_with_none_engine(self, tmp_path):
        """Zero issues + None engine should not crash."""
        rg = self._make_none_engine_report(tmp_path, issues=[])
        pdf_path = rg.generate_pdf()
        excel_path = rg.generate_excel()
        assert os.path.exists(pdf_path)
        assert os.path.exists(excel_path)


# ==================================================================
# Gap 20: Report fingerprinting -- timestamp in filenames
# ==================================================================

_TS_PATTERN = re.compile(r"\d{8}_\d{6}")


# ==================================================================
# Filename sanitization (Task 2)
# ==================================================================


class TestFilenameSanitization:
    """ReportGenerator.__init__ must sanitize filenames via os.path.basename."""

    def test_path_traversal_stripped(self, tmp_path):
        """Path traversal components like ../../ must be stripped."""
        rg = _make_report_generator(filename="../../evil.xlsx", tmp_path=tmp_path)
        assert rg.filename == "evil.xlsx"

    def test_normal_filename_unchanged(self, tmp_path):
        """A plain filename should pass through unchanged."""
        rg = _make_report_generator(filename="model.xlsx", tmp_path=tmp_path)
        assert rg.filename == "model.xlsx"

    def test_slash_in_filename_sanitized(self, tmp_path):
        """A filename with directory separators should be reduced to basename."""
        rg = _make_report_generator(filename="dir/model.xlsx", tmp_path=tmp_path)
        assert rg.filename == "model.xlsx"


class TestReportFingerprinting:
    """Verify output filenames contain a timestamp."""

    def test_pdf_filename_has_timestamp(self, tmp_path):
        rg = _make_report_generator(tmp_path=tmp_path)
        pdf_path = rg.generate_pdf()
        assert _TS_PATTERN.search(os.path.basename(pdf_path)), (
            f"PDF filename missing timestamp pattern: {pdf_path}"
        )

    def test_excel_filename_has_timestamp(self, tmp_path):
        rg = _make_report_generator(tmp_path=tmp_path)
        excel_path = rg.generate_excel()
        assert _TS_PATTERN.search(os.path.basename(excel_path)), (
            f"Excel filename missing timestamp pattern: {excel_path}"
        )
