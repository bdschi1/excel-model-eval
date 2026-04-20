"""
Tests for src/ingestion.py -- ModelIngestor

Verifies:
1. Successful xlsx ingestion (dual-state: values + formulas)
2. File-not-found handling
3. Corrupted file handling
4. sheets_values / sheets_formulas populated correctly
5. get_ingestion_report() output format
6. Empty workbook handling
"""

from __future__ import annotations

import pathlib
import sys

import openpyxl
import pandas as pd

# Ensure repo root is on path
REPO_ROOT = pathlib.Path(__file__).parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from src.ingestion import ModelIngestor

# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------

def _make_workbook(tmp_path, sheets_data, filename="test.xlsx"):
    """
    Create a real xlsx file from a dict of {sheet_name: [[row1], [row2], ...]}.
    Returns the file path.
    """
    wb = openpyxl.Workbook()
    first = True
    for sheet_name, rows in sheets_data.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        for row_idx, row in enumerate(rows, 1):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
    path = str(tmp_path / filename)
    wb.save(path)
    return path


# ==================================================================
# Tests
# ==================================================================


class TestIngestionSuccess:
    """Happy-path ingestion tests."""

    def test_ingest_returns_true_on_valid_xlsx(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Revenue": [["Item", "2023", "2024"], ["Sales", 100, 120]],
        })
        ingestor = ModelIngestor(path)
        result = ingestor.ingest()
        assert result is True

    def test_sheets_values_populated(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Sheet1": [[10, 20], [30, 40]],
        })
        ingestor = ModelIngestor(path)
        ingestor.ingest()
        assert "Sheet1" in ingestor.sheets_values
        df = ingestor.sheets_values["Sheet1"]
        assert isinstance(df, pd.DataFrame)
        assert df.shape == (2, 2)

    def test_sheets_formulas_populated(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Sheet1": [[10, 20], [30, "=A1+A2"]],
        })
        ingestor = ModelIngestor(path)
        ingestor.ingest()
        assert "Sheet1" in ingestor.sheets_formulas
        df = ingestor.sheets_formulas["Sheet1"]
        assert isinstance(df, pd.DataFrame)
        # The formula cell should be a string starting with '='
        formula_val = df.iloc[1, 1]
        assert isinstance(formula_val, str) and formula_val.startswith("=")

    def test_values_layer_returns_computed_values(self, tmp_path):
        """data_only=True should return computed values (None for unsaved)."""
        path = _make_workbook(tmp_path, {
            "Sheet1": [[10, 20], [30, "=A1+B1"]],
        })
        ingestor = ModelIngestor(path)
        ingestor.ingest()
        # openpyxl data_only=True on a freshly-saved file returns None
        # for formula cells (no cached value). This is expected behavior.
        val = ingestor.sheets_values["Sheet1"].iloc[1, 1]
        # Should be None (not yet computed by Excel) or numeric
        assert val is None or isinstance(val, (int, float))

    def test_multiple_sheets(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Income": [["Revenue", 1000]],
            "Balance": [["Assets", 5000]],
            "Cash": [["Operating", 200]],
        })
        ingestor = ModelIngestor(path)
        ingestor.ingest()
        assert len(ingestor.sheets_values) == 3
        assert set(ingestor.sheets_values.keys()) == {"Income", "Balance", "Cash"}
        assert len(ingestor.sheets_formulas) == 3


class TestIngestionFailures:
    """Error handling tests."""

    def test_file_not_found_returns_false(self):
        ingestor = ModelIngestor("/nonexistent/path/model.xlsx")
        result = ingestor.ingest()
        assert result is False

    def test_corrupted_file_returns_false(self, tmp_path):
        bad_path = str(tmp_path / "corrupt.xlsx")
        with open(bad_path, "wb") as f:
            f.write(b"this is not a valid xlsx file at all")
        ingestor = ModelIngestor(bad_path)
        result = ingestor.ingest()
        assert result is False

    def test_empty_workbook(self, tmp_path):
        """An empty workbook (no data in any cell) should ingest without error."""
        wb = openpyxl.Workbook()
        path = str(tmp_path / "empty.xlsx")
        wb.save(path)
        ingestor = ModelIngestor(path)
        result = ingestor.ingest()
        assert result is True
        # Should have one sheet (default "Sheet")
        assert len(ingestor.sheets_values) == 1


class TestIngestionReport:
    """get_ingestion_report() structure tests."""

    def test_report_structure(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Summary": [["Test", 1]],
        })
        ingestor = ModelIngestor(path)
        ingestor.ingest()
        report = ingestor.get_ingestion_report()

        assert "total_sheets" in report
        assert "sheet_names" in report
        assert "errors" in report
        assert "status" in report

    def test_report_total_sheets(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "A": [[1]], "B": [[2]], "C": [[3]],
        })
        ingestor = ModelIngestor(path)
        ingestor.ingest()
        report = ingestor.get_ingestion_report()
        assert report["total_sheets"] == 3

    def test_report_sheet_names(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Revenue": [[1]], "COGS": [[2]],
        })
        ingestor = ModelIngestor(path)
        ingestor.ingest()
        report = ingestor.get_ingestion_report()
        assert set(report["sheet_names"]) == {"Revenue", "COGS"}

    def test_report_status_success_no_errors(self, tmp_path):
        path = _make_workbook(tmp_path, {"OK": [[1]]})
        ingestor = ModelIngestor(path)
        ingestor.ingest()
        report = ingestor.get_ingestion_report()
        assert report["status"] == "Success"
        assert report["errors"] == []

    def test_report_before_ingestion(self):
        """Report before ingest() should show zero sheets."""
        ingestor = ModelIngestor("dummy.xlsx")
        report = ingestor.get_ingestion_report()
        assert report["total_sheets"] == 0
        assert report["sheet_names"] == []


class TestIngestionFilename:
    """Verify filename property is set correctly."""

    def test_filename_extracted(self, tmp_path):
        path = _make_workbook(tmp_path, {"S": [[1]]}, filename="model_v2.xlsx")
        ingestor = ModelIngestor(path)
        assert ingestor.filename == "model_v2.xlsx"


# ==================================================================
# Exception narrowing (Task 5)
# ==================================================================


class TestExceptionNarrowing:
    """Verify ingest() handles specific exceptions gracefully."""

    def test_corrupted_zip_returns_false(self, tmp_path):
        """Writing garbage bytes to a .xlsx file should cause ingest() to
        return False without raising an exception."""
        bad_path = str(tmp_path / "garbage.xlsx")
        with open(bad_path, "wb") as f:
            f.write(b"\x00\x01\x02\x03GARBAGE_NOT_A_ZIP")
        ingestor = ModelIngestor(bad_path)
        result = ingestor.ingest()
        assert result is False
        # Should not raise; errors should be recorded
        assert len(ingestor.load_errors) > 0

    def test_timeout_returns_false_with_error(self, tmp_path):
        """When _load_with_timeout raises IngestionTimeout, ingest() should
        return False and record a timeout message in load_errors."""
        from unittest.mock import patch

        from src.ingestion import IngestionTimeout  # noqa: F811

        # Create a valid xlsx so we pass the file-exists and magic-bytes checks
        path = _make_workbook(tmp_path, {"S": [[1]]}, filename="timeout_test.xlsx")
        ingestor = ModelIngestor(path)

        with patch.object(
            ingestor,
            "_load_with_timeout",
            side_effect=IngestionTimeout("Ingestion timed out after 120s"),
        ):
            result = ingestor.ingest()

        assert result is False
        timeout_errors = [e for e in ingestor.load_errors if "TIMEOUT" in e.upper()]
        assert len(timeout_errors) >= 1
